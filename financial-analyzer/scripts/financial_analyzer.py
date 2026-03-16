#!/usr/bin/env python3
"""
附注优先的财务分析主脚本。
"""

import argparse
import collections
import datetime
import hashlib
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass


ENGINE_VERSION = "3.1.0"

TOPIC_PATTERNS = {
    "audit": ["核数师", "审计", "审计意见", "保留意见", "无法表示意见", "否定意见"],
    "debt": ["借贷", "借款", "债券", "贷款", "融资", "票据", "债项"],
    "cash": ["现金", "现金流", "银行存款", "受限资金", "货币资金"],
    "profitability": ["收入", "利润", "溢利", "毛利", "EBITDA", "收益"],
    "liquidity": ["流动", "短期", "到期", "偿债", "再融资", "期限结构"],
    "investment_property": ["投资物业", "公允价值", "估值", "资本化率", "物业估值"],
    "restricted_assets": ["受限资金", "受限资产", "抵押", "质押", "冻结资金", "保证金"],
    "risk_management": ["风险", "掉期", "对冲", "汇率", "利率", "契约"],
    "lgfv_features": ["LGFV", "城投", "国资", "政府补助", "专项债", "资本公积注入", "化债", "置换债"],
    "external_guarantees": ["对外担保", "担保网络", "被担保单位", "互保", "反担保"],
    "governance": ["董事会", "管治", "薪酬", "合规", "股东"],
    "sustainability": ["可持续", "碳", "环保", "绿色融资", "ESG"],
    "tax": ["税项", "递延税项", "所得税", "税率"],
}

RISK_RULES = [
    {"name": "extreme_audit_issue", "severity": "extreme", "patterns": ["无法表示意见", "否定意见"]},
    {"name": "high_audit_issue", "severity": "high", "patterns": ["保留意见", "持续经营"]},
    {"name": "litigation_or_default", "severity": "high", "patterns": ["违约", "逾期", "诉讼", "仲裁"]},
    {"name": "liquidity_pressure", "severity": "high", "patterns": ["一年内到期", "再融资", "短期债务", "流动性风险"]},
    {"name": "asset_impairment", "severity": "high", "patterns": ["减值", "减值准备", "公允价值减少"]},
    {"name": "restricted_cash", "severity": "medium", "patterns": ["受限资金", "受限制", "质押"]},
    {"name": "interest_capitalization", "severity": "medium", "patterns": ["资本化之借贷支出", "利息资本化", "资本化利息"]},
    {"name": "guarantee_exposure", "severity": "medium", "patterns": ["担保", "反担保"]},
    {"name": "fx_or_rate_exposure", "severity": "medium", "patterns": ["汇率", "利率掉期", "货币掉期", "净投资对冲"]},
]

STOPWORDS = {
    "目录",
    "附注",
    "章节",
    "项目",
    "年度",
    "报告",
    "综合",
    "本集团",
    "本公司",
    "第",
    "项",
    "表",
}

SEVERITY_SCORE = {
    "extreme": 100,
    "high": 70,
    "medium": 45,
    "low": 20,
}

RISK_LABELS = {
    "extreme_audit_issue": "极端审计意见风险",
    "high_audit_issue": "审计或持续经营提示",
    "litigation_or_default": "违约或诉讼风险",
    "liquidity_pressure": "流动性与再融资压力",
    "asset_impairment": "资产减值或公允价值下行压力",
    "restricted_cash": "受限资金占用风险",
    "interest_capitalization": "资本化利息口径风险",
    "guarantee_exposure": "担保敞口风险",
    "fx_or_rate_exposure": "汇率或利率敞口风险",
}

FORMAL_TOPIC_MODULES = {
    "investment_property": "投资性房地产",
    "restricted_assets": "受限资产与抵押",
    "lgfv_features": "城投平台特征",
    "inventory_quality": "存货质量",
    "external_guarantees": "对外担保",
}

CASE_TOPIC_ALLOWLIST = {
    "henglong": {"investment_property"},
    "country_garden": {"restricted_assets"},
    "hanghai": {"lgfv_features", "external_guarantees"},
}


def read_text(path):
    with open(path, "r", encoding="utf-8") as handle:
        return handle.read()


def now_iso():
    return datetime.datetime.now().astimezone().isoformat(timespec="seconds")


def write_json(path, payload):
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def write_jsonl(path, rows):
    with open(path, "w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def read_lines(path):
    with open(path, "r", encoding="utf-8") as handle:
        return handle.read().splitlines()


def md5_file(path):
    digest = hashlib.md5()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_text(text):
    text = text.replace("\ufeff", "")
    text = re.sub(r"\r\n?", "\n", text)
    return text


def clean_line(line):
    line = re.sub(r"!\[[^\]]*\]\([^)]+\)", " ", line)
    line = re.sub(r"<[^>]+>", " ", line)
    line = re.sub(r"\s+", " ", line)
    return line.strip()


def shorten(text, limit=120):
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def fallback_company_name(stem):
    cleaned = re.sub(r"(20\d{2}).*$", "", stem)
    cleaned = re.sub(r"[（(].*?[)）]", "", cleaned)
    return cleaned.strip("-_ ")


def extract_company_name(text, fallback_name):
    patterns = [
        r"^\s*([^\n]{2,40}?(?:股份有限公司|集团股份有限公司|有限公司|集团有限公司))(?:（|\(|\s|$)",
        r"公司名称[：:]\s*([^\n]{2,50})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, re.MULTILINE)
        if match:
            return match.group(1).lstrip("# ").strip()
    return fallback_company_name(fallback_name).lstrip("# ").strip()


def extract_report_period(text, fallback_name):
    stem_match = re.search(r"(20\d{2})", fallback_name)
    if stem_match:
        return stem_match.group(1)

    patterns = [
        r"(20\d{2})年度报告",
        r"(20\d{2})年年度报告",
        r"(20\d{2})年报",
        r"截至(20\d{2})年12月31日止年度",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    return ""


def detect_currency(text):
    if "港币" in text or "港元" in text:
        return "HKD"
    if "美元" in text or "USD" in text:
        return "USD"
    return "CNY"


def classify_report(text, md_path):
    lower_name = md_path.name.lower()
    report_type = "a_share_full_report"
    basis = []

    if "香港财务报告准则" in text or "港币" in text or "hkfrs" in lower_name or "年報" in md_path.name:
        report_type = "hong_kong_full_report"
        basis.append("检测到港币/HKFRS/年報字样")
    elif "交易商协会" in text or ("合并资产负债表" not in text and "财务报表附注" not in text):
        report_type = "nfmii_brief_report"
        basis.append("未检测到完整报表与附注结构")
    else:
        basis.append("检测到完整财务报表与附注结构")

    if re.search(r"无法表示意见", text):
        audit_opinion = "无法表示意见"
    elif re.search(r"否定意见", text):
        audit_opinion = "否定意见"
    elif re.search(r"我们认为[\s\S]{0,200}真实而中肯地反映", text):
        audit_opinion = "标准无保留"
    elif re.search(r"保留意见的基础|发表保留意见", text):
        audit_opinion = "保留意见"
    elif re.search(r"强调事项", text):
        audit_opinion = "带强调事项"
    else:
        audit_opinion = "未识别"

    return {
        "report_type": report_type,
        "audit_opinion": audit_opinion,
        "classification_basis": basis,
    }


def resolve_run_dir(run_dir_arg):
    run_dir = Path(run_dir_arg).resolve()
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def load_notes_workfile(notes_workfile_path, total_lines):
    notes_workfile = Path(notes_workfile_path).resolve()
    if not notes_workfile.exists():
        raise FileNotFoundError(f"notes_workfile 不存在: {notes_workfile}")

    with open(notes_workfile, "r", encoding="utf-8") as handle:
        payload = json.load(handle)

    required_fields = [
        "notes_start_line",
        "notes_end_line",
        "locator_evidence",
        "notes_catalog",
    ]
    missing_fields = [field for field in required_fields if field not in payload]
    if missing_fields:
        raise ValueError(f"notes_workfile 缺少字段: {', '.join(missing_fields)}")

    notes_start_line = payload["notes_start_line"]
    notes_end_line = payload["notes_end_line"]
    notes_catalog = payload["notes_catalog"]
    locator_evidence = payload["locator_evidence"]

    if not isinstance(notes_start_line, int) or not isinstance(notes_end_line, int):
        raise ValueError("notes_start_line 和 notes_end_line 必须为整数")
    if notes_start_line < 1 or notes_end_line < notes_start_line or notes_end_line > total_lines:
        raise ValueError("附注区间无效")
    if not isinstance(locator_evidence, list) or not locator_evidence:
        raise ValueError("locator_evidence 不能为空")
    if not isinstance(notes_catalog, list) or not notes_catalog:
        raise ValueError("notes_catalog 不能为空")

    normalized_catalog = []
    required_note_fields = ["note_no", "chapter_title", "start_line", "end_line", "evidence"]
    previous_end_line = notes_start_line - 1
    for index, note in enumerate(notes_catalog, start=1):
        missing_note_fields = [field for field in required_note_fields if field not in note]
        if missing_note_fields:
            raise ValueError(f"notes_catalog 第 {index} 项缺少字段: {', '.join(missing_note_fields)}")

        chapter_title = str(note["chapter_title"]).strip()
        start_line = note["start_line"]
        end_line = note["end_line"]
        evidence = note["evidence"]

        if not chapter_title:
            raise ValueError(f"notes_catalog 第 {index} 项 chapter_title 不能为空")
        if not isinstance(start_line, int) or not isinstance(end_line, int):
            raise ValueError(f"notes_catalog 第 {index} 项边界必须为整数")
        if start_line < notes_start_line or end_line > notes_end_line or end_line < start_line:
            raise ValueError(f"notes_catalog 第 {index} 项边界超出附注区间")
        if start_line <= previous_end_line:
            raise ValueError(f"notes_catalog 第 {index} 项边界未按顺序排列")
        if not isinstance(evidence, list) or not evidence:
            raise ValueError(f"notes_catalog 第 {index} 项 evidence 不能为空")

        normalized_catalog.append({
            "note_no": str(note["note_no"]).strip(),
            "chapter_title": chapter_title,
            "start_line": start_line,
            "end_line": end_line,
            "evidence": evidence,
        })
        previous_end_line = end_line

    return {
        "path": str(notes_workfile),
        "notes_start_line": notes_start_line,
        "notes_end_line": notes_end_line,
        "locator_evidence": locator_evidence,
        "notes_catalog": normalized_catalog,
    }


def split_sentences(text):
    text = clean_line(text)
    parts = re.split(r"[。！？!?；;\n]+", text)
    return [item.strip() for item in parts if len(item.strip()) >= 8]


def summarize_chapter(text):
    sentences = split_sentences(text)
    if not sentences:
        return "章节内容较少，保留原始证据以待后续补充。"
    return "；".join(sentences[:2])


def numeric_value(raw):
    raw = raw.replace(",", "").strip()
    raw = raw.rstrip("%")
    try:
        return float(raw)
    except ValueError:
        return None


def detect_unit(line):
    candidates = [
        "港币百万元",
        "人民币百万元",
        "亿港元",
        "亿元",
        "亿",
        "百万",
        "万元",
        "元",
        "%",
        "倍",
    ]
    for candidate in candidates:
        if candidate in line:
            return candidate
    return ""


def extract_numeric_data(text, limit=12):
    entries = []
    seen_lines = set()
    for raw_line in text.splitlines():
        line = clean_line(raw_line)
        if not line or line in seen_lines:
            continue
        if not re.search(r"\d", line):
            continue
        matches = re.findall(r"(?<!\d)(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?%?", line)
        if not matches:
            continue

        seen_lines.add(line)
        value = numeric_value(matches[0])
        prefix = line.split(matches[0], 1)[0].strip(" ：:，,.-")
        label = shorten(prefix or line, limit=40)
        entries.append({
            "label": label or "数值片段",
            "value": value,
            "raw_value": matches[0],
            "unit": detect_unit(line),
            "evidence": shorten(line, limit=160),
        })

    entries.sort(
        key=lambda item: (
            0 if item["value"] is None else -abs(item["value"]),
            item["label"],
        )
    )
    return entries[:limit]


def extract_title_tokens(title):
    tokens = []
    for token in re.findall(r"[A-Za-z]{2,}|[\u4e00-\u9fff]{2,12}", title):
        candidate = token.strip()
        if candidate in STOPWORDS:
            continue
        if candidate.startswith("财务报表"):
            continue
        if candidate not in tokens:
            tokens.append(candidate)
    return tokens


def infer_topics(title, text):
    topics = []
    content = f"{title}\n{text}"

    for topic_name, patterns in TOPIC_PATTERNS.items():
        if any(pattern in content for pattern in patterns):
            topics.append(topic_name)

    for token in extract_title_tokens(title):
        if token not in topics:
            topics.append(token)

    if not topics:
        topics.append("general")

    return topics[:8]


def detect_case_key(company_name, input_file):
    content = f"{company_name} {input_file}"
    if "恒隆" in content:
        return "henglong"
    if "碧桂园" in content:
        return "country_garden"
    if "杭海新城" in content:
        return "hanghai"
    return ""


def find_evidence_lines(text, patterns, limit=2):
    evidences = []
    for raw_line in text.splitlines():
        line = clean_line(raw_line)
        if not line:
            continue
        if any(pattern in line for pattern in patterns):
            evidences.append(shorten(line, limit=160))
        if len(evidences) >= limit:
            break
    return evidences


def build_findings(title, text, topic_tags, numeric_data):
    findings = []
    candidate_sentences = split_sentences(text)

    for sentence in candidate_sentences:
        if len(findings) >= 4:
            break
        if re.search(r"\d", sentence) or any(tag in sentence for tag in topic_tags if len(tag) >= 2):
            findings.append({
                "finding_name": shorten(f"{title}关键信息", limit=40),
                "detail": shorten(sentence, limit=160),
                "topics": topic_tags[:3],
                "confidence": "medium",
            })

    if not findings:
        findings.append({
            "finding_name": shorten(f"{title}章节概览", limit=40),
            "detail": summarize_chapter(text),
            "topics": topic_tags[:3],
            "confidence": "medium",
        })

    if numeric_data and len(findings) < 5:
        first_numeric = numeric_data[0]
        findings.append({
            "finding_name": shorten(f"{title}数值提要", limit=40),
            "detail": first_numeric["evidence"],
            "topics": topic_tags[:3],
            "confidence": "medium",
        })

    return findings[:5]


def build_anomalies(text):
    anomalies = []
    for rule in RISK_RULES:
        evidence_lines = find_evidence_lines(text, rule["patterns"], limit=2)
        if not evidence_lines:
            continue
        anomalies.append({
            "signal_name": rule["name"],
            "severity": rule["severity"],
            "evidence": evidence_lines,
            "impact_hint": f"检测到 {', '.join(rule['patterns'][:2])} 相关表述",
        })
    return anomalies


def build_evidence(title, text, numeric_data, anomalies):
    evidence = []
    summary = summarize_chapter(text)
    if summary:
        evidence.append({
            "source": title,
            "type": "summary",
            "content": summary,
        })

    for item in numeric_data[:4]:
        evidence.append({
            "source": title,
            "type": "numeric_data",
            "content": item["evidence"],
        })

    for anomaly in anomalies[:3]:
        for line in anomaly["evidence"]:
            evidence.append({
                "source": title,
                "type": "risk_signal",
                "content": line,
            })

    return evidence[:8]


def build_chapter_record(index, note_entry, markdown_lines, report_context, locator_evidence):
    text = "\n".join(markdown_lines[note_entry["start_line"] - 1:note_entry["end_line"]]).strip()
    cleaned_text = normalize_text(text)
    chapter_title = f"{note_entry['note_no']} {note_entry['chapter_title']}".strip()
    topic_tags = infer_topics(chapter_title, cleaned_text)
    numeric_data = extract_numeric_data(cleaned_text)
    anomalies = build_anomalies(cleaned_text)
    findings = build_findings(chapter_title, cleaned_text, topic_tags, numeric_data)
    evidence = build_evidence(chapter_title, cleaned_text, numeric_data, anomalies)
    status = "completed" if cleaned_text else "empty"

    return {
        "chapter_no": index,
        "chapter_title": chapter_title,
        "status": status,
        "summary": summarize_chapter(cleaned_text),
        "attributes": {
            "chapter_type": "notes_chapter",
            "note_no": note_entry["note_no"],
            "note_scope": "notes_only",
            "locator_evidence": locator_evidence,
            "topic_tags": topic_tags,
            "line_span": {
                "start": note_entry["start_line"],
                "end": note_entry["end_line"],
            },
            "line_count": note_entry["end_line"] - note_entry["start_line"] + 1,
            "char_count": len(cleaned_text),
            "table_count": cleaned_text.count("<table>"),
            "image_count": cleaned_text.count("![]("),
            "report_type": report_context["report_type"],
        },
        "numeric_data": numeric_data,
        "findings": findings,
        "anomalies": anomalies,
        "evidence": evidence,
        "extensions": {
            "raw_title_tokens": extract_title_tokens(chapter_title),
            "section_evidence": note_entry["evidence"],
            "dynamic_topics": [tag for tag in topic_tags if tag not in TOPIC_PATTERNS],
        },
    }


def group_focus_candidates(chapter_records):
    grouped = {}
    for record in chapter_records:
        base_topics = record["attributes"]["topic_tags"][:3] or ["general"]
        for anomaly in record["anomalies"]:
            key = anomaly["signal_name"]
            grouped.setdefault(key, {
                "focus_name": key,
                "score": 0,
                "severity": anomaly["severity"],
                "evidence_chapters": set(),
                "related_topics": set(base_topics),
                "evidence_samples": [],
            })
            item = grouped[key]
            item["score"] += SEVERITY_SCORE.get(anomaly["severity"], 20)
            item["evidence_chapters"].add(record["chapter_no"])
            item["related_topics"].update(base_topics)
            item["evidence_samples"].extend(anomaly["evidence"][:1])

        if record["numeric_data"]:
            key = base_topics[0]
            grouped.setdefault(key, {
                "focus_name": key,
                "score": 0,
                "severity": "low",
                "evidence_chapters": set(),
                "related_topics": set(base_topics),
                "evidence_samples": [],
            })
            item = grouped[key]
            item["score"] += min(len(record["numeric_data"]) * 4, 20)
            item["evidence_chapters"].add(record["chapter_no"])
            item["related_topics"].update(base_topics)
            item["evidence_samples"].extend(
                [entry["evidence"] for entry in record["numeric_data"][:1]]
            )

    return grouped


def build_focus_list(chapter_records):
    candidates = list(group_focus_candidates(chapter_records).values())
    candidates.sort(key=lambda item: (-item["score"], item["focus_name"]))

    focus_list = []
    for candidate in candidates[:6]:
        evidence_chapters = sorted(candidate["evidence_chapters"])
        focus_list.append({
            "focus_name": candidate["focus_name"],
            "why_selected": shorten(
                "基于章节证据密度与风险信号强度自动选出，需在综合分析时优先展开。",
                limit=80,
            ),
            "evidence_chapters": evidence_chapters,
            "focus_attributes": {
                "severity": candidate["severity"],
                "score": candidate["score"],
                "evidence_count": len(candidate["evidence_samples"]),
            },
            "related_topics": sorted(candidate["related_topics"]),
            "knowledge_gap": (
                "待确认是否需要沉淀为正式规则"
                if candidate["focus_name"] not in TOPIC_PATTERNS
                else "已有主题，可继续补充案例证据"
            ),
            "impact_scope": sorted(candidate["related_topics"])[:3],
        })
    return focus_list


def aggregate_topic_results(chapter_records):
    topic_map = collections.OrderedDict()
    for record in chapter_records:
        for topic in record["attributes"]["topic_tags"]:
            if topic not in topic_map:
                topic_map[topic] = {
                    "summary_parts": [],
                    "chapters": [],
                    "numeric_highlights": [],
                    "anomalies": [],
                }
            bucket = topic_map[topic]
            bucket["summary_parts"].append(record["summary"])
            bucket["chapters"].append(record["chapter_no"])
            bucket["numeric_highlights"].extend(record["numeric_data"][:2])
            bucket["anomalies"].extend(record["anomalies"][:2])

    results = {}
    for topic, bucket in topic_map.items():
        results[topic] = {
            "summary": shorten("；".join(bucket["summary_parts"][:3]), limit=200),
            "attributes": {
                "chapter_count": len(sorted(set(bucket["chapters"]))),
                "chapter_refs": sorted(set(bucket["chapters"])),
            },
            "evidence": [item["evidence"] for item in bucket["numeric_highlights"][:4]],
            "extensions": {
                "risk_signals": bucket["anomalies"][:3],
            },
        }
    return results


def build_final_data(report_context, chapter_records, focus_list):
    topic_results = aggregate_topic_results(chapter_records)
    conclusions = []

    conclusions.append(
        f"报告类型识别为 {report_context['report_type']}，审计意见为 {report_context['audit_opinion']}。"
    )
    if focus_list:
        conclusions.append(
            f"本次动态重点共 {len(focus_list)} 项，优先关注 {focus_list[0]['focus_name']}。"
        )
    conclusions.append(
        f"附注章节记录共生成 {len(chapter_records)} 条，覆盖章节 {chapter_records[0]['chapter_no']} 至 {chapter_records[-1]['chapter_no']}。"
    )

    return {
        "entity_profile": {
            "company_name": report_context["company_name"],
            "report_period": report_context["report_period"],
            "currency": report_context["currency"],
            "report_type": report_context["report_type"],
            "audit_opinion": report_context["audit_opinion"],
            "input_file": report_context["input_file"],
        },
        "key_conclusions": conclusions,
        "topic_results": topic_results,
        "extensions": {
            "focus_count": len(focus_list),
            "chapter_count": len(chapter_records),
            "classification_basis": report_context["classification_basis"],
        },
    }


def chapter_context_text(record):
    parts = [record["summary"]]
    parts.extend(item["content"] for item in record.get("evidence", []))
    parts.extend(item["detail"] for item in record.get("findings", []))
    parts.extend(item["evidence"] for item in record.get("numeric_data", []))
    for anomaly in record.get("anomalies", []):
        parts.extend(anomaly.get("evidence", []))
    return "\n".join(item for item in parts if item)


def infer_industry_tag(chapter_records):
    topic_tags = set()
    for record in chapter_records:
        topic_tags.update(record["attributes"].get("topic_tags", []))

    if "investment_property" in topic_tags:
        return "property"
    if "lgfv_features" in topic_tags or any("城投" in tag for tag in topic_tags):
        return "lgfv"
    return "general"


def normalize_amount_value(raw_value, evidence, currency, prefer_large_unit=False):
    if raw_value is None:
        return None, ""

    evidence = evidence or ""
    currency_prefix = {
        "HKD": "HKD",
        "USD": "USD",
        "CNY": "CNY",
    }.get(currency, currency or "")

    if "%" in evidence:
        return raw_value, "%"

    if "倍" in evidence:
        return raw_value, "x"

    if "亿元" in evidence or "亿港元" in evidence or "亿美元" in evidence:
        return raw_value, f"{currency_prefix}_100m"

    if "百万元" in evidence or "百萬元" in evidence:
        if prefer_large_unit:
            return round(raw_value / 100.0, 2), f"{currency_prefix}_100m"
        return raw_value, f"{currency_prefix}_mn"

    if prefer_large_unit and raw_value >= 1000:
        return round(raw_value / 100.0, 2), f"{currency_prefix}_100m"

    if raw_value >= 1000:
        return raw_value, f"{currency_prefix}_mn"

    return raw_value, ""


def make_metric_fact(
    metric_code,
    label,
    value,
    unit,
    period,
    source_status,
    evidence_refs=None,
    comparison=None,
    benchmark=None,
    risk_level="unknown",
):
    return {
        "metric_code": metric_code,
        "label": label,
        "value": value,
        "unit": unit,
        "period": period,
        "comparison": comparison,
        "benchmark": benchmark,
        "risk_level": risk_level,
        "source_status": source_status,
        "evidence_refs": evidence_refs or [],
    }


def make_table_row(
    row_code,
    label,
    values,
    source_status,
    evidence_refs=None,
    commentary="",
):
    return {
        "row_code": row_code,
        "label": label,
        "values": values,
        "commentary": commentary,
        "source_status": source_status,
        "evidence_refs": evidence_refs or [],
    }


def make_manifest_item(sheet_name, module_key, module_type, required, enabled, title, empty_state):
    return {
        "sheet_name": sheet_name,
        "module_key": module_key,
        "module_type": module_type,
        "required": required,
        "enabled": enabled,
        "title": title,
        "empty_state": empty_state,
    }


def risk_level_from_value(metric_code, value):
    if value is None:
        return "unknown"
    if metric_code == "net_debt_to_equity":
        if value >= 60:
            return "high"
        if value >= 40:
            return "medium"
        return "low"
    if metric_code == "cash_to_short_term_debt":
        if value < 1:
            return "high"
        if value < 1.5:
            return "medium"
        return "low"
    return "medium"


def build_soul_export_payload(report_context, notes_work, run_dir, final_data, focus_list, chapter_records):
    period = report_context["report_period"]
    period_end = f"{period}-12-31" if period else ""
    currency = report_context["currency"]
    case_key = detect_case_key(report_context["company_name"], report_context["input_file"])
    source_artifacts = {
        "run_manifest": str(run_dir / "run_manifest.json"),
        "chapter_records": str(run_dir / "chapter_records.jsonl"),
        "focus_list": str(run_dir / "focus_list.json"),
        "final_data": str(run_dir / "final_data.json"),
        "analysis_report": str(run_dir / "analysis_report.md"),
        "financial_output": str(run_dir / "financial_output.xlsx"),
        "notes_workfile": notes_work["path"],
    }

    evidence_index = []
    next_evidence_id = 1

    def add_evidence(field_path, sheet_name, excerpt, confidence="medium", record=None, source_document="annual_report_notes"):
        nonlocal next_evidence_id
        clean_excerpt = shorten(clean_line(excerpt), limit=160)
        if not clean_excerpt:
            return []

        evidence_id = f"EVD-{next_evidence_id:04d}"
        next_evidence_id += 1
        note_no = None
        line_span = {"start": None, "end": None}
        chapter_no = None
        chapter_title = ""
        if record:
            note_no = record["attributes"].get("note_no")
            line_span = record["attributes"].get("line_span", line_span)
            chapter_no = record.get("chapter_no")
            chapter_title = record.get("chapter_title", "")

        evidence_index.append({
            "evidence_id": evidence_id,
            "field_path": field_path,
            "sheet_name": sheet_name,
            "excerpt": clean_excerpt,
            "source_document": source_document,
            "chapter_no": chapter_no,
            "chapter_title": chapter_title,
            "note_no": note_no,
            "line_span": line_span,
            "confidence": confidence,
        })
        return [evidence_id]

    topic_records = collections.defaultdict(list)
    for record in chapter_records:
        for tag in record["attributes"].get("topic_tags", []):
            topic_records[tag].append(record)

    def records_for_tags(*tags):
        results = []
        seen = set()
        for tag in tags:
            for record in topic_records.get(tag, []):
                if record["chapter_no"] not in seen:
                    results.append(record)
                    seen.add(record["chapter_no"])
        return results

    def pick_record(records, title_keyword=None):
        if title_keyword:
            for record in records:
                if title_keyword in record["chapter_title"]:
                    return record
        return records[0] if records else None

    def search_record(record, pattern, flags=0):
        if not record:
            return None
        match = re.search(pattern, chapter_context_text(record), flags)
        if not match:
            return None
        raw_value = numeric_value(match.group(1))
        return {
            "value": raw_value,
            "excerpt": shorten(clean_line(match.group(0)), limit=160),
            "record": record,
        }

    def search_records(records, pattern, flags=0, title_keyword=None):
        for record in records:
            if title_keyword and title_keyword not in record["chapter_title"]:
                continue
            found = search_record(record, pattern, flags)
            if found:
                return found
        return None

    debt_records = records_for_tags("debt", "liquidity", "risk_management")
    cash_records = records_for_tags("cash")
    profitability_records = records_for_tags("profitability")
    investment_property_records = records_for_tags("investment_property")
    restricted_asset_records = records_for_tags("restricted_assets", "cash")
    lgfv_records = records_for_tags("lgfv_features")
    guarantee_records = records_for_tags("external_guarantees")

    debt_record = pick_record(debt_records, "借贷")
    cash_record = pick_record(cash_records, "现金")
    profitability_record = profitability_records[0] if profitability_records else None
    bond_record = (
        pick_record(debt_records, "债券")
        or pick_record(debt_records, "发债")
        or pick_record(debt_records, "票据")
        or debt_record
    )

    debt_total = search_record(debt_record, r"银行贷款及其他借贷总额\s*\(?([\d,]+(?:\.\d+)?)", re.S)
    short_term_debt = search_record(debt_record, r"列入流动负债下一年内到期款项\s*\(?([\d,]+(?:\.\d+)?)", re.S)
    net_debt = search_record(cash_record, r"净债项\s*\(?([\d,]+(?:\.\d+)?)", re.S)
    unused_credit = search_record(debt_record, r"未动用的银行承诺信贷额度[^\d]*港币([\d.]+)亿元", re.S)
    mtn_remainder = search_record(debt_record, r"等值港币([\d.]+)亿元", re.S)
    covenant_linked = search_record(debt_record, r"账面值港币([\d.]+)亿元[^。]*借贷受财务契约限制", re.S)
    cash_equiv = search_record(cash_record, r"现金及现金等价物\s*\(?([\d,]+(?:\.\d+)?)", re.S)
    cash_deposits = search_record(cash_record, r"现金及银行存款\s*\(?([\d,]+(?:\.\d+)?)", re.S)
    net_debt_to_equity = search_record(profitability_record, r"净债项股权比率\s*([\d.]+)%", re.S)
    bank_rate = re.search(r"按([\d.]+)%至([\d.]+)%", chapter_context_text(debt_record) if debt_record else "", re.S)
    bond_coupon = re.search(r"票面年利率为([\d.]+)%至([\d.]+)%", chapter_context_text(debt_record) if debt_record else "", re.S)
    debt_compliance = "compliant" if debt_record and "完全遵守" in chapter_context_text(debt_record) else "unknown"
    restricted_cash = search_records(restricted_asset_records, r"受限资金[=：:\s]*\(?([\d,]+(?:\.\d+)?)", re.S)
    free_cash = search_records(restricted_asset_records, r"自由现金[=：:\s]*\(?([\d,]+(?:\.\d+)?)", re.S)
    pledged_asset = search_records(restricted_asset_records, r"(?:抵押）|抵押物)[^\d]{0,16}([\d,]+(?:\.\d+)?)", re.S)
    guarantee_balance = search_records(guarantee_records, r"担保(?:金额|余额)?[^\d]{0,20}([\d,]+(?:\.\d+)?)", re.S)
    government_subsidy = search_records(lgfv_records, r"政府补助[^\d]{0,20}([\d,]+(?:\.\d+)?)", re.S)
    capital_injection = search_records(lgfv_records, r"资本公积(?:注入)?[^\d]{0,20}([\d,]+(?:\.\d+)?)", re.S)
    offshore_bond = search_records(debt_records + lgfv_records, r"境外发债(?:结构)?[^\d]{0,20}([\d,]+(?:\.\d+)?)", re.S)
    bond_balance = search_records(debt_records, r"应付债券[^\d]{0,20}([\d,]+(?:\.\d+)?)", re.S)
    total_debt_fallback = search_records(debt_records + lgfv_records, r"有息债务合计[^\d~]{0,8}~?([\d,]+(?:\.\d+)?)", re.S)

    if debt_total is None and total_debt_fallback:
        debt_total = total_debt_fallback
    if short_term_debt is None:
        short_term_debt = search_records(
            debt_records,
            r"(?:一年内到期[^\\d]{0,20}|短期借款[^\\d]{0,20})([\d,]+(?:\.\d+)?)",
            re.S,
        )
    if cash_equiv is None and free_cash:
        cash_equiv = free_cash

    overview_highlights = []
    locator_excerpt = "；".join(item.get("excerpt", "") for item in notes_work["locator_evidence"][:2] if item.get("excerpt"))
    locator_refs = add_evidence(
        "overview.report_highlights.locator",
        "00_overview",
        locator_excerpt or "附注定位成功",
        confidence="high",
        source_document="notes_workfile",
    )
    overview_highlights.append({
        "title": "附注定位",
        "detail": f"已定位附注区间 {notes_work['notes_catalog'][0]['note_no']} - {notes_work['notes_catalog'][-1]['note_no']}",
        "evidence_refs": locator_refs,
    })
    overview_highlights.append({
        "title": "附注覆盖",
        "detail": f"当前覆盖 {len(chapter_records)} 个附注章节",
        "evidence_refs": [],
    })

    seen_risks = set()
    key_risks = []
    for record in chapter_records:
        for anomaly in record["anomalies"]:
            signal_name = anomaly["signal_name"]
            if signal_name in seen_risks:
                continue
            seen_risks.add(signal_name)
            risk_refs = []
            for excerpt in anomaly.get("evidence", [])[:1]:
                risk_refs.extend(
                    add_evidence(
                        f"overview.key_risks.{signal_name}",
                        "00_overview",
                        excerpt,
                        confidence=anomaly["severity"],
                        record=record,
                    )
                )
            key_risks.append({
                "risk_code": signal_name,
                "label": RISK_LABELS.get(signal_name, signal_name),
                "risk_level": anomaly["severity"],
                "description": anomaly["impact_hint"],
                "evidence_refs": risk_refs,
            })
            if len(key_risks) >= 5:
                break
        if len(key_risks) >= 5:
            break

    overview = {
        "executive_summary": final_data["key_conclusions"][:5],
        "key_risks": key_risks,
        "rating_snapshot": [],
        "report_highlights": overview_highlights,
    }

    total_debt_value = None
    total_debt_unit = ""
    total_debt_refs = []
    if debt_total:
        total_debt_value, total_debt_unit = normalize_amount_value(
            debt_total["value"],
            debt_total["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        total_debt_refs = add_evidence(
            "debt_profile.totals.total_interest_bearing_debt",
            "03_debt_profile",
            debt_total["excerpt"],
            record=debt_total["record"],
        )

    cash_value = None
    cash_unit = ""
    cash_refs = []
    if cash_equiv:
        cash_value, cash_unit = normalize_amount_value(
            cash_equiv["value"],
            cash_equiv["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        cash_refs = add_evidence(
            "liquidity_and_covenants.cash_metrics.cash_and_cash_equiv",
            "04_liquidity_and_covenants",
            cash_equiv["excerpt"],
            record=cash_equiv["record"],
        )

    short_term_value = None
    short_term_unit = ""
    short_term_refs = []
    if short_term_debt:
        short_term_value, short_term_unit = normalize_amount_value(
            short_term_debt["value"],
            short_term_debt["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        short_term_refs = add_evidence(
            "debt_profile.maturity_buckets.within_1y",
            "03_debt_profile",
            short_term_debt["excerpt"],
            record=short_term_debt["record"],
        )

    long_term_value = None
    if total_debt_value is not None and short_term_value is not None:
        long_term_value = round(total_debt_value - short_term_value, 2)

    cash_short_ratio = None
    if cash_value is not None and short_term_value not in (None, 0):
        cash_short_ratio = round(cash_value / short_term_value, 2)

    unused_credit_value = None
    unused_credit_refs = []
    if unused_credit:
        unused_credit_value, unused_credit_unit = normalize_amount_value(
            unused_credit["value"],
            unused_credit["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        unused_credit_refs = add_evidence(
            "liquidity_and_covenants.credit_lines.unused_committed_facility",
            "04_liquidity_and_covenants",
            unused_credit["excerpt"],
            record=unused_credit["record"],
        )
    else:
        unused_credit_unit = ""

    mtn_value = None
    mtn_refs = []
    bond_metric = mtn_remainder or bond_balance or offshore_bond
    if bond_metric:
        mtn_value, mtn_unit = normalize_amount_value(
            bond_metric["value"],
            bond_metric["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        mtn_refs = add_evidence(
            "optional_modules.bond_detail.bonds.mtn_program",
            "05_bond_detail",
            bond_metric["excerpt"],
            record=bond_metric["record"],
        )
    else:
        mtn_unit = ""

    covenant_value = None
    covenant_refs = []
    if covenant_linked:
        covenant_value, covenant_unit = normalize_amount_value(
            covenant_linked["value"],
            covenant_linked["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        covenant_refs = add_evidence(
            "liquidity_and_covenants.covenants.covenant_linked_debt",
            "04_liquidity_and_covenants",
            covenant_linked["excerpt"],
            record=covenant_linked["record"],
        )
    else:
        covenant_unit = ""

    restricted_cash_value = None
    restricted_cash_unit = ""
    restricted_cash_refs = []
    if restricted_cash:
        restricted_cash_value, restricted_cash_unit = normalize_amount_value(
            restricted_cash["value"],
            restricted_cash["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        restricted_cash_refs = add_evidence(
            "liquidity_and_covenants.restricted_assets.restricted_cash",
            "04_liquidity_and_covenants",
            restricted_cash["excerpt"],
            record=restricted_cash["record"],
        )

    pledged_asset_value = None
    pledged_asset_unit = ""
    pledged_asset_refs = []
    if pledged_asset:
        pledged_asset_value, pledged_asset_unit = normalize_amount_value(
            pledged_asset["value"],
            pledged_asset["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        pledged_asset_refs = add_evidence(
            "liquidity_and_covenants.restricted_assets.pledged_assets",
            "04_liquidity_and_covenants",
            pledged_asset["excerpt"],
            record=pledged_asset["record"],
        )

    guarantee_value = None
    guarantee_unit = ""
    guarantee_refs = []
    if guarantee_balance:
        guarantee_value, guarantee_unit = normalize_amount_value(
            guarantee_balance["value"],
            guarantee_balance["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        guarantee_refs = add_evidence(
            "optional_modules.topic_external_guarantees.sections.guarantee_balance",
            "08_topic_external_guarantees",
            guarantee_balance["excerpt"],
            record=guarantee_balance["record"],
        )

    subsidy_value = None
    subsidy_unit = ""
    subsidy_refs = []
    if government_subsidy:
        subsidy_value, subsidy_unit = normalize_amount_value(
            government_subsidy["value"],
            government_subsidy["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        subsidy_refs = add_evidence(
            "optional_modules.topic_lgfv_features.sections.government_subsidy",
            "08_topic_lgfv_features",
            government_subsidy["excerpt"],
            record=government_subsidy["record"],
        )

    capital_value = None
    capital_unit = ""
    capital_refs = []
    if capital_injection:
        capital_value, capital_unit = normalize_amount_value(
            capital_injection["value"],
            capital_injection["excerpt"],
            currency,
            prefer_large_unit=True,
        )
        capital_refs = add_evidence(
            "optional_modules.topic_lgfv_features.sections.capital_injection",
            "08_topic_lgfv_features",
            capital_injection["excerpt"],
            record=capital_injection["record"],
        )

    debt_profile = {
        "totals": [
            make_metric_fact(
                "total_interest_bearing_debt",
                "有息债务总额",
                total_debt_value,
                total_debt_unit,
                period_end,
                "direct" if debt_total else "manual_needed",
                total_debt_refs,
                risk_level="medium" if total_debt_value is not None else "unknown",
            ),
            make_metric_fact(
                "net_debt",
                "净债务",
                *normalize_amount_value(net_debt["value"], net_debt["excerpt"], currency, prefer_large_unit=True)
                if net_debt else (None, ""),
                period_end,
                "direct" if net_debt else "manual_needed",
                add_evidence(
                    "debt_profile.totals.net_debt",
                    "03_debt_profile",
                    net_debt["excerpt"],
                    record=net_debt["record"],
                ) if net_debt else [],
                risk_level="medium" if net_debt else "unknown",
            ),
        ],
        "maturity_buckets": [
            make_table_row(
                "within_1y",
                "一年内到期",
                [{"period": period_end, "value": short_term_value, "unit": short_term_unit}],
                "direct" if short_term_debt else "manual_needed",
                short_term_refs,
            ),
            make_table_row(
                "after_1y",
                "一年后到期",
                [{"period": period_end, "value": long_term_value, "unit": short_term_unit or total_debt_unit}],
                "derived" if long_term_value is not None else "manual_needed",
                short_term_refs + total_debt_refs if long_term_value is not None else [],
            ),
        ],
        "financing_mix": [
            make_table_row(
                "bond",
                "债券/票据",
                [{"period": period_end, "value": mtn_value, "unit": mtn_unit}],
                "direct" if mtn_remainder else "manual_needed",
                mtn_refs,
            ),
        ],
        "rate_profile": [
            make_table_row(
                "bank_loan_rate_range",
                "银行贷款利率区间",
                [{"period": period_end, "value": f"{bank_rate.group(1)}%~{bank_rate.group(2)}%" if bank_rate else None, "unit": "text"}],
                "direct" if bank_rate else "manual_needed",
                add_evidence(
                    "debt_profile.rate_profile.bank_loan_rate_range",
                    "03_debt_profile",
                    bank_rate.group(0),
                    record=debt_record,
                ) if bank_rate else [],
            ),
            make_table_row(
                "bond_coupon_range",
                "债券票面利率区间",
                [{"period": period_end, "value": f"{bond_coupon.group(1)}%~{bond_coupon.group(2)}%" if bond_coupon else None, "unit": "text"}],
                "direct" if bond_coupon else "manual_needed",
                add_evidence(
                    "debt_profile.rate_profile.bond_coupon_range",
                    "03_debt_profile",
                    bond_coupon.group(0),
                    record=debt_record,
                ) if bond_coupon else [],
            ),
        ],
        "debt_comments": [
            {
                "label": "契约约束借款",
                "detail": f"{covenant_value}{covenant_unit}" if covenant_value is not None else "",
                "source_status": "direct" if covenant_linked else "manual_needed",
                "evidence_refs": covenant_refs,
            },
        ],
    }

    kpi_dashboard = {
        "periods": [period],
        "sections": [
            {
                "category": "leverage",
                "metrics": [
                    make_metric_fact(
                        "net_debt_to_equity",
                        "净债务/权益",
                        net_debt_to_equity["value"] if net_debt_to_equity else None,
                        "%",
                        period_end,
                        "direct" if net_debt_to_equity else "manual_needed",
                        add_evidence(
                            "kpi_dashboard.sections.leverage.net_debt_to_equity",
                            "01_kpi_dashboard",
                            net_debt_to_equity["excerpt"],
                            record=net_debt_to_equity["record"],
                        ) if net_debt_to_equity else [],
                        risk_level=risk_level_from_value("net_debt_to_equity", net_debt_to_equity["value"] if net_debt_to_equity else None),
                    ),
                    make_metric_fact(
                        "total_interest_bearing_debt",
                        "有息债务总额",
                        total_debt_value,
                        total_debt_unit,
                        period_end,
                        "direct" if debt_total else "manual_needed",
                        total_debt_refs,
                    ),
                ],
            },
            {
                "category": "debt_service",
                "metrics": [
                    make_metric_fact(
                        "cash_to_short_term_debt",
                        "现金短债比",
                        cash_short_ratio,
                        "x",
                        period_end,
                        "derived" if cash_short_ratio is not None else "manual_needed",
                        cash_refs + short_term_refs if cash_short_ratio is not None else [],
                        risk_level=risk_level_from_value("cash_to_short_term_debt", cash_short_ratio),
                    ),
                    make_metric_fact(
                        "unused_committed_facility",
                        "未动用授信",
                        unused_credit_value,
                        unused_credit_unit,
                        period_end,
                        "direct" if unused_credit else "manual_needed",
                        unused_credit_refs,
                    ),
                ],
            },
            {
                "category": "profitability",
                "metrics": [
                    make_metric_fact(
                        "operating_profit",
                        "营业利润",
                        None,
                        "",
                        period_end,
                        "manual_needed",
                    ),
                ],
            },
            {
                "category": "cashflow",
                "metrics": [
                    make_metric_fact(
                        "operating_cash_flow",
                        "经营现金流净额",
                        None,
                        "",
                        period_end,
                        "manual_needed",
                    ),
                ],
            },
        ],
    }

    liquidity_and_covenants = {
        "cash_metrics": [
            make_metric_fact(
                "cash_and_cash_equiv",
                "现金及现金等价物",
                cash_value,
                cash_unit,
                period_end,
                "direct" if cash_equiv else "manual_needed",
                cash_refs,
            ),
            make_metric_fact(
                "cash_and_bank_deposits",
                "现金及银行存款",
                *normalize_amount_value(cash_deposits["value"], cash_deposits["excerpt"], currency, prefer_large_unit=True)
                if cash_deposits else (None, ""),
                period_end,
                "direct" if cash_deposits else "manual_needed",
                add_evidence(
                    "liquidity_and_covenants.cash_metrics.cash_and_bank_deposits",
                    "04_liquidity_and_covenants",
                    cash_deposits["excerpt"],
                    record=cash_deposits["record"],
                ) if cash_deposits else [],
            ),
        ],
        "credit_lines": [
            make_table_row(
                "unused_bank_committed_facility",
                "未动用银行承诺授信",
                [{"period": period_end, "value": unused_credit_value, "unit": unused_credit_unit}],
                "direct" if unused_credit else "manual_needed",
                unused_credit_refs,
            ),
            make_table_row(
                "mtn_program_headroom",
                "中票/票据计划可发行余额",
                [{"period": period_end, "value": mtn_value, "unit": mtn_unit}],
                "direct" if bond_metric else "manual_needed",
                mtn_refs,
            ),
        ],
        "restricted_assets": [
            make_table_row(
                "restricted_cash",
                "受限资金",
                [{"period": period_end, "value": restricted_cash_value, "unit": restricted_cash_unit}],
                "direct" if restricted_cash else "manual_needed",
                restricted_cash_refs,
            ),
            make_table_row(
                "pledged_assets",
                "抵押/质押资产",
                [{"period": period_end, "value": pledged_asset_value, "unit": pledged_asset_unit}],
                "direct" if pledged_asset else "manual_needed",
                pledged_asset_refs,
            ),
        ],
        "covenants": [
            {
                "covenant_code": "debt_covenant_compliance",
                "label": "财务契约合规",
                "status": debt_compliance,
                "restricted_debt": covenant_value,
                "unit": covenant_unit,
                "source_status": "direct" if covenant_linked else "manual_needed",
                "evidence_refs": covenant_refs,
            },
        ],
        "liquidity_observations": [
            {
                "label": "流动性信号",
                "detail": "存在一年内到期债务与再融资管理要求。" if short_term_debt else "流动性关键信息待补充。",
                "source_status": "derived" if short_term_debt else "manual_needed",
                "evidence_refs": short_term_refs,
            },
            {
                "label": "受限资产提示",
                "detail": "存在受限资金或抵押资产，需要结合自由现金判断可动用流动性。" if restricted_cash or pledged_asset else "受限资产信息待补充。",
                "source_status": "derived" if restricted_cash or pledged_asset else "manual_needed",
                "evidence_refs": restricted_cash_refs + pledged_asset_refs,
            },
        ],
    }

    financial_summary = {
        "unit_label": f"{currency}_100m" if currency else "",
        "statements": {
            "balance_sheet": [],
            "income_statement": [],
            "cash_flow": [],
        },
        "coverage_note": "当前 notes-only 链路无法稳定生成摘要三表，已保留结构待后续标准化报表抽取补足。",
    }

    optional_modules = []
    bond_context = chapter_context_text(bond_record) if bond_record else ""
    if bond_metric or re.search(r"债券|票据|中期票据|美元债|发债", bond_context):
        optional_modules.append({
            "module_key": "bond_detail",
            "sheet_name": "05_bond_detail",
            "module_type": "optional",
            "title": "债券明细",
            "payload": {
                "bonds": [
                    {
                        "instrument_name": "中期票据计划余额" if mtn_remainder else "债券/票据",
                        "balance": mtn_value,
                        "unit": mtn_unit,
                        "coupon_range": f"{bond_coupon.group(1)}%~{bond_coupon.group(2)}%" if bond_coupon else "",
                        "issue_date": "",
                        "maturity_date": "",
                        "terms": "",
                        "guarantee": "",
                        "source_status": "direct" if bond_metric else "manual_needed",
                        "evidence_refs": mtn_refs or (
                            add_evidence(
                                "optional_modules.bond_detail.bonds.fallback",
                                "05_bond_detail",
                                bond_record["summary"],
                                record=bond_record,
                            ) if bond_record else []
                        ),
                    }
                ],
            },
        })

    allowed_topic_keys = CASE_TOPIC_ALLOWLIST.get(case_key)
    for topic_key, display_name in FORMAL_TOPIC_MODULES.items():
        if allowed_topic_keys is not None and topic_key not in allowed_topic_keys:
            continue
        if topic_key not in topic_records and topic_key not in final_data["topic_results"]:
            continue
        topic_record = pick_record(topic_records.get(topic_key, []))
        topic_payload = final_data["topic_results"].get(topic_key, {})
        topic_summary = topic_payload.get("summary") or (topic_record["summary"] if topic_record else "")
        topic_refs = add_evidence(
            f"optional_modules.topic_{topic_key}.summary",
            f"08_topic_{topic_key}",
            topic_summary,
            record=topic_record,
        ) if topic_summary and topic_record else []
        section_items = []
        if topic_record:
            for numeric_item in topic_record.get("numeric_data", [])[:2]:
                section_items.append({
                    "label": numeric_item["label"],
                    "value": numeric_item["value"],
                    "unit": numeric_item["unit"],
                    "source_status": "direct",
                    "evidence_refs": add_evidence(
                        f"optional_modules.topic_{topic_key}.sections.numeric",
                        f"08_topic_{topic_key}",
                        numeric_item["evidence"],
                        record=topic_record,
                    ),
                })
        if topic_key == "restricted_assets":
            if restricted_cash_value is not None:
                section_items.append({
                    "label": "受限资金",
                    "value": restricted_cash_value,
                    "unit": restricted_cash_unit,
                    "source_status": "direct",
                    "evidence_refs": restricted_cash_refs,
                })
            if pledged_asset_value is not None:
                section_items.append({
                    "label": "抵押/质押资产",
                    "value": pledged_asset_value,
                    "unit": pledged_asset_unit,
                    "source_status": "direct",
                    "evidence_refs": pledged_asset_refs,
                })
        if topic_key == "lgfv_features":
            if subsidy_value is not None:
                section_items.append({
                    "label": "政府补助",
                    "value": subsidy_value,
                    "unit": subsidy_unit,
                    "source_status": "direct",
                    "evidence_refs": subsidy_refs,
                })
            if capital_value is not None:
                section_items.append({
                    "label": "资本公积注入",
                    "value": capital_value,
                    "unit": capital_unit,
                    "source_status": "direct",
                    "evidence_refs": capital_refs,
                })
        if topic_key == "external_guarantees" and guarantee_value is not None:
            section_items.append({
                "label": "对外担保余额",
                "value": guarantee_value,
                "unit": guarantee_unit,
                "source_status": "direct",
                "evidence_refs": guarantee_refs,
            })
        optional_modules.append({
            "module_key": f"topic_{topic_key}",
            "sheet_name": f"08_topic_{topic_key}",
            "module_type": "topic",
            "title": display_name,
            "payload": {
                "topic_key": topic_key,
                "display_name": display_name,
                "summary": topic_summary,
                "summary_evidence_refs": topic_refs,
                "sections": [
                    {
                        "section_title": "案例要点",
                        "items": section_items,
                    }
                ],
                "ext_fields": {
                    "risk_signals": topic_payload.get("extensions", {}).get("risk_signals", []),
                },
            },
        })

    module_manifest = [
        make_manifest_item("00_overview", "overview", "fixed", True, True, "概览", "概览为空时保留执行摘要与风险骨架"),
        make_manifest_item("01_kpi_dashboard", "kpi_dashboard", "fixed", True, True, "KPI 面板", "无法稳定识别的指标使用 null + source_status"),
        make_manifest_item("02_financial_summary", "financial_summary", "fixed", True, True, "财务摘要", "notes-only 场景允许空载荷"),
        make_manifest_item("03_debt_profile", "debt_profile", "fixed", True, True, "债务画像", "债务拆分不足时保留表头骨架"),
        make_manifest_item("04_liquidity_and_covenants", "liquidity_and_covenants", "fixed", True, True, "流动性与契约", "缺少受限资产时保留空数组"),
    ]
    for module in optional_modules:
        module_manifest.append(
            make_manifest_item(
                module["sheet_name"],
                module["module_key"],
                module["module_type"],
                False,
                True,
                module["title"],
                "按适用性启用，可为空但不默认输出",
            )
        )
    module_manifest.append(
        make_manifest_item("99_evidence_index", "evidence_index", "fixed", True, True, "证据索引", "所有 evidence_refs 必须落地到该表")
    )

    return {
        "contract_version": "soul_export_v1",
        "template_version": "soul_v1_1_alpha",
        "generated_at": now_iso(),
        "entity_profile": {
            "company_name": final_data["entity_profile"]["company_name"],
            "report_period": final_data["entity_profile"]["report_period"],
            "currency": final_data["entity_profile"]["currency"],
            "report_type": final_data["entity_profile"]["report_type"],
            "audit_opinion": final_data["entity_profile"]["audit_opinion"],
            "industry_tag": infer_industry_tag(chapter_records),
            "input_file": final_data["entity_profile"]["input_file"],
        },
        "source_artifacts": source_artifacts,
        "module_manifest": module_manifest,
        "overview": overview,
        "kpi_dashboard": kpi_dashboard,
        "financial_summary": financial_summary,
        "debt_profile": debt_profile,
        "liquidity_and_covenants": liquidity_and_covenants,
        "optional_modules": optional_modules,
        "evidence_index": evidence_index,
    }


def report_scope_hint(topic_tags):
    if "investment_property" in topic_tags:
        return "适用于投资性房地产占比较高的地产企业"
    if "restricted_assets" in topic_tags:
        return "适用于受限资金、抵押资产较多的地产或高杠杆主体"
    if "lgfv_features" in topic_tags:
        return "适用于城投平台或政府支持特征显著的主体"
    if "external_guarantees" in topic_tags:
        return "适用于存在明显担保网络或互保关系的主体"
    if "debt" in topic_tags or "liquidity" in topic_tags:
        return "适用于存在明显债务融资结构的发行人"
    if "audit" in topic_tags:
        return "适用于所有需要审计意见判断的完整年报"
    return "适用于案例相近的完整年报分析"


def build_pending_updates(chapter_records):
    items = []
    seen_titles = set()

    for record in chapter_records:
        for topic in record["extensions"]["dynamic_topics"]:
            if topic in seen_titles:
                continue
            seen_titles.add(topic)
            items.append({
                "type": "topic_candidate",
                "title": topic,
                "proposal": f"将章节主题“{topic}”作为开放主题标签候选，等待更多案例验证。",
                "source": f"chapter_{record['chapter_no']}",
                "evidence": record["evidence"][:2],
                "applicable_scope": report_scope_hint(record["attributes"]["topic_tags"]),
                "status": "candidate",
                "introduced_in": ENGINE_VERSION,
                "confidence": "medium",
            })

        for item in record["numeric_data"][:2]:
            label = item["label"]
            if len(label) < 4 or label in seen_titles:
                continue
            seen_titles.add(label)
            items.append({
                "type": "field_candidate",
                "title": label,
                "proposal": f"新增可选扩展字段“{label}”，保留为案例级扩展载荷。",
                "source": f"chapter_{record['chapter_no']}",
                "evidence": [{"source": record["chapter_title"], "type": "numeric_data", "content": item["evidence"]}],
                "applicable_scope": report_scope_hint(record["attributes"]["topic_tags"]),
                "status": "candidate",
                "introduced_in": ENGINE_VERSION,
                "confidence": "low",
            })

        for anomaly in record["anomalies"]:
            if anomaly["signal_name"] in seen_titles:
                continue
            seen_titles.add(anomaly["signal_name"])
            items.append({
                "type": "rule_candidate",
                "title": anomaly["signal_name"],
                "proposal": f"补充“{anomaly['signal_name']}”识别规则，并明确证据模式与适用范围。",
                "source": f"chapter_{record['chapter_no']}",
                "evidence": [{"source": record["chapter_title"], "type": "risk_signal", "content": line} for line in anomaly["evidence"]],
                "applicable_scope": report_scope_hint(record["attributes"]["topic_tags"]),
                "status": "candidate",
                "introduced_in": ENGINE_VERSION,
                "confidence": anomaly["severity"],
            })

        if len(items) >= 12:
            break

    return {
        "metadata": {
            "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
            "required_fields": [
                "source",
                "evidence",
                "applicable_scope",
                "status",
                "introduced_in",
                "confidence",
            ],
            "governance_rule": "缺少元数据的候选项不能直接进入正式知识库。",
        },
        "items": items[:12],
    }


def build_report_markdown(report_context, focus_list, final_data, pending_updates, chapter_records):
    lines = [
        f"# {report_context['company_name']} {report_context['report_period']} 年报分析报告",
        "",
        "## 运行概览",
        f"- 报告类型：{report_context['report_type']}",
        f"- 审计意见：{report_context['audit_opinion']}",
        f"- 币种：{report_context['currency']}",
        f"- 附注章节记录数：{len(chapter_records)}",
        "",
        "## 动态重点",
    ]

    for focus in focus_list:
        lines.append(
            f"- `{focus['focus_name']}`：{focus['why_selected']}（章节：{', '.join(str(item) for item in focus['evidence_chapters'])}）"
        )

    lines.extend([
        "",
        "## 关键结论",
    ])

    for conclusion in final_data["key_conclusions"]:
        lines.append(f"- {conclusion}")

    lines.extend([
        "",
        "## 章节速览",
    ])

    for record in chapter_records[:12]:
        lines.append(
            f"- 附注第{record['attributes']['note_no']}章 `{record['chapter_title']}`：{record['summary']}"
        )

    lines.extend([
        "",
        "## 待固化更新",
    ])

    for item in pending_updates["items"][:10]:
        lines.append(f"- `{item['type']}` / `{item['title']}`：{item['proposal']}")

    lines.append("")
    return "\n".join(lines)


def autosize_sheet(sheet):
    for column in sheet.columns:
        max_length = 0
        letter = column[0].column_letter
        for cell in column:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[letter].width = min(max_length + 2, 50)


def build_workbook(report_context, focus_list, final_data, pending_updates, chapter_records, output_path):
    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    summary_sheet = workbook.active
    summary_sheet.title = "summary"
    summary_rows = [
        ("company_name", report_context["company_name"]),
        ("report_period", report_context["report_period"]),
        ("report_type", report_context["report_type"]),
        ("audit_opinion", report_context["audit_opinion"]),
        ("currency", report_context["currency"]),
        ("chapter_count", len(chapter_records)),
        ("focus_count", len(focus_list)),
    ]
    for row_index, (key, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=row_index, column=1, value=key)
        summary_sheet.cell(row=row_index, column=2, value=value)

    focus_sheet = workbook.create_sheet("focus")
    focus_headers = ["focus_name", "severity", "score", "evidence_chapters", "related_topics", "knowledge_gap"]
    for col_index, header in enumerate(focus_headers, start=1):
        cell = focus_sheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, focus in enumerate(focus_list, start=2):
        focus_sheet.cell(row=row_index, column=1, value=focus["focus_name"])
        focus_sheet.cell(row=row_index, column=2, value=focus["focus_attributes"]["severity"])
        focus_sheet.cell(row=row_index, column=3, value=focus["focus_attributes"]["score"])
        focus_sheet.cell(row=row_index, column=4, value=", ".join(str(item) for item in focus["evidence_chapters"]))
        focus_sheet.cell(row=row_index, column=5, value=", ".join(focus["related_topics"]))
        focus_sheet.cell(row=row_index, column=6, value=focus["knowledge_gap"])

    chapter_sheet = workbook.create_sheet("chapters")
    chapter_headers = ["chapter_no", "chapter_title", "status", "topics", "summary", "anomaly_count", "numeric_count"]
    for col_index, header in enumerate(chapter_headers, start=1):
        cell = chapter_sheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, record in enumerate(chapter_records, start=2):
        chapter_sheet.cell(row=row_index, column=1, value=record["chapter_no"])
        chapter_sheet.cell(row=row_index, column=2, value=record["chapter_title"])
        chapter_sheet.cell(row=row_index, column=3, value=record["status"])
        chapter_sheet.cell(row=row_index, column=4, value=", ".join(record["attributes"]["topic_tags"]))
        chapter_sheet.cell(row=row_index, column=5, value=record["summary"])
        chapter_sheet.cell(row=row_index, column=6, value=len(record["anomalies"]))
        chapter_sheet.cell(row=row_index, column=7, value=len(record["numeric_data"]))

    topic_sheet = workbook.create_sheet("topic_results")
    topic_headers = ["topic", "chapter_count", "chapter_refs", "summary"]
    for col_index, header in enumerate(topic_headers, start=1):
        cell = topic_sheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, (topic, payload) in enumerate(final_data["topic_results"].items(), start=2):
        topic_sheet.cell(row=row_index, column=1, value=topic)
        topic_sheet.cell(row=row_index, column=2, value=payload["attributes"]["chapter_count"])
        topic_sheet.cell(row=row_index, column=3, value=", ".join(str(item) for item in payload["attributes"]["chapter_refs"]))
        topic_sheet.cell(row=row_index, column=4, value=payload["summary"])

    update_sheet = workbook.create_sheet("pending_updates")
    update_headers = ["type", "title", "status", "confidence", "applicable_scope", "proposal"]
    for col_index, header in enumerate(update_headers, start=1):
        cell = update_sheet.cell(row=1, column=col_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, item in enumerate(pending_updates["items"], start=2):
        update_sheet.cell(row=row_index, column=1, value=item["type"])
        update_sheet.cell(row=row_index, column=2, value=item["title"])
        update_sheet.cell(row=row_index, column=3, value=item["status"])
        update_sheet.cell(row=row_index, column=4, value=item["confidence"])
        update_sheet.cell(row=row_index, column=5, value=item["applicable_scope"])
        update_sheet.cell(row=row_index, column=6, value=item["proposal"])

    for sheet in workbook.worksheets:
        autosize_sheet(sheet)

    workbook.save(output_path)


def build_manifest(md_path, notes_work, run_dir, report_context, chapter_records, focus_list, pending_updates):
    return {
        "engine_version": ENGINE_VERSION,
        "status": "success",
        "failure_reason": "",
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "input": {
            "md_path": str(md_path),
            "md5": md5_file(md_path),
            "file_size": md_path.stat().st_size,
            "notes_workfile": notes_work["path"],
        },
        "entity": {
            "company_name": report_context["company_name"],
            "report_period": report_context["report_period"],
            "report_type": report_context["report_type"],
            "audit_opinion": report_context["audit_opinion"],
        },
        "notes_locator": {
            "status": "success",
            "start_line": notes_work["notes_start_line"],
            "end_line": notes_work["notes_end_line"],
            "locator_evidence": notes_work["locator_evidence"],
        },
        "notes_catalog_summary": {
            "note_chapter_count": len(notes_work["notes_catalog"]),
            "first_note": notes_work["notes_catalog"][0]["note_no"],
            "last_note": notes_work["notes_catalog"][-1]["note_no"],
        },
        "artifacts": {
            "run_manifest": str(run_dir / "run_manifest.json"),
            "chapter_records": str(run_dir / "chapter_records.jsonl"),
            "focus_list": str(run_dir / "focus_list.json"),
            "final_data": str(run_dir / "final_data.json"),
            "soul_export_payload": str(run_dir / "soul_export_payload.json"),
            "pending_updates": str(run_dir / "pending_updates.json"),
            "analysis_report": str(run_dir / "analysis_report.md"),
            "financial_output": str(run_dir / "financial_output.xlsx"),
        },
        "counts": {
            "chapter_records": len(chapter_records),
            "focus_items": len(focus_list),
            "pending_updates": len(pending_updates["items"]),
        },
    }


def build_failure_manifest(md_path, notes_workfile_path, run_dir, report_context, failure_reason, details):
    return {
        "engine_version": ENGINE_VERSION,
        "status": "failed",
        "failure_reason": failure_reason,
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "input": {
            "md_path": str(md_path),
            "md5": md5_file(md_path),
            "file_size": md_path.stat().st_size,
            "notes_workfile": str(Path(notes_workfile_path).resolve()),
        },
        "entity": {
            "company_name": report_context["company_name"],
            "report_period": report_context["report_period"],
            "report_type": report_context["report_type"],
            "audit_opinion": report_context["audit_opinion"],
        },
        "notes_locator": {
            "status": "failed",
            "start_line": None,
            "end_line": None,
            "locator_evidence": [],
        },
        "notes_catalog_summary": {
            "note_chapter_count": 0,
            "first_note": "",
            "last_note": "",
        },
        "artifacts": {
            "run_manifest": str(run_dir / "run_manifest.json"),
        },
        "details": details,
    }


def fail_with_manifest(md_path, notes_workfile_path, run_dir, report_context, failure_reason, details):
    manifest = build_failure_manifest(
        md_path,
        notes_workfile_path,
        run_dir,
        report_context,
        failure_reason,
        details,
    )
    write_json(run_dir / "run_manifest.json", manifest)
    raise SystemExit(1)


def main():
    parser = argparse.ArgumentParser(description="附注优先的财务分析主脚本")
    parser.add_argument("--md", required=True, help="输入 Markdown 路径")
    parser.add_argument("--notes-workfile", required=True, help="附注工作文件路径")
    parser.add_argument("--run-dir", required=True, help="显式指定运行目录")
    args = parser.parse_args()

    md_path = Path(args.md).resolve()
    if not md_path.exists():
        raise FileNotFoundError(f"Markdown 文件不存在: {md_path}")

    run_dir = resolve_run_dir(args.run_dir)
    stable_excel_path = run_dir / "financial_output.xlsx"

    print(f"[INFO] 输入文件: {md_path}")
    print(f"[INFO] 运行目录: {run_dir}")

    raw_text = normalize_text(read_text(md_path))
    markdown_lines = read_lines(md_path)
    company_name = extract_company_name(raw_text, md_path.stem)
    report_period = extract_report_period(raw_text, md_path.stem)
    report_context = {
        "company_name": company_name,
        "report_period": report_period,
        "currency": detect_currency(raw_text),
        "input_file": str(md_path),
    }
    report_context.update(classify_report(raw_text, md_path))

    try:
        notes_work = load_notes_workfile(args.notes_workfile, len(markdown_lines))
    except FileNotFoundError as exc:
        fail_with_manifest(md_path, args.notes_workfile, run_dir, report_context, "notes_workfile_missing", str(exc))
    except ValueError as exc:
        fail_with_manifest(md_path, args.notes_workfile, run_dir, report_context, "notes_workfile_invalid", str(exc))

    chapter_records = [
        build_chapter_record(
            index,
            note_entry,
            markdown_lines,
            report_context,
            notes_work["locator_evidence"],
        )
        for index, note_entry in enumerate(notes_work["notes_catalog"], start=1)
    ]
    if not chapter_records:
        fail_with_manifest(md_path, args.notes_workfile, run_dir, report_context, "notes_catalog_empty", "notes_catalog 为空")

    focus_list = build_focus_list(chapter_records)
    final_data = build_final_data(report_context, chapter_records, focus_list)
    pending_updates = build_pending_updates(chapter_records)
    soul_export_payload = build_soul_export_payload(
        report_context,
        notes_work,
        run_dir,
        final_data,
        focus_list,
        chapter_records,
    )
    analysis_report = build_report_markdown(
        report_context,
        focus_list,
        final_data,
        pending_updates,
        chapter_records,
    )

    manifest = build_manifest(
        md_path,
        notes_work,
        run_dir,
        report_context,
        chapter_records,
        focus_list,
        pending_updates,
    )

    write_json(run_dir / "run_manifest.json", manifest)
    write_jsonl(run_dir / "chapter_records.jsonl", chapter_records)
    write_json(run_dir / "focus_list.json", focus_list)
    write_json(run_dir / "final_data.json", final_data)
    write_json(run_dir / "soul_export_payload.json", soul_export_payload)
    write_json(run_dir / "pending_updates.json", pending_updates)

    with open(run_dir / "analysis_report.md", "w", encoding="utf-8") as handle:
        handle.write(analysis_report)

    build_workbook(
        report_context,
        focus_list,
        final_data,
        pending_updates,
        chapter_records,
        stable_excel_path,
    )

    print(f"[OK] 章节记录: {len(chapter_records)}")
    print(f"[OK] 动态重点: {len(focus_list)}")
    print(f"[OK] 待固化更新: {len(pending_updates['items'])}")
    print(f"✅ 成功: 产物已生成 -> {run_dir}")


if __name__ == "__main__":
    main()
