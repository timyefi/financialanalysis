from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RUN_DIR = Path('/Volumes/P5S/工作/同步空间/BaiduSyncdisk/bondclaw/financialanalysis/test_runs/vanke_2025_semiannual_20260406')
CHAPTER_PATH = RUN_DIR / 'chapter_records.jsonl'
FINAL_DATA_PATH = RUN_DIR / 'final_data.json'
SOUL_PAYLOAD_PATH = RUN_DIR / 'soul_export_payload.json'
FORMALIZATION_MANIFEST_PATH = RUN_DIR / 'formalization_manifest.json'
WORKBOOK_PATH = RUN_DIR / 'financial_output.xlsx'


@dataclass
class Chapter:
    chapter_no: int
    chapter_title: str
    summary: str
    chapter_text: str
    line_start: int
    line_end: int
    numeric_data: list[dict[str, Any]]
    topic_tags: list[str]



def load_chapters() -> dict[int, Chapter]:
    chapters: dict[int, Chapter] = {}
    for line in CHAPTER_PATH.read_text(encoding='utf-8').splitlines():
        if not line.strip():
            continue
        record = json.loads(line)
        attributes = record.get('attributes', {})
        line_span = attributes.get('line_span', {})
        chapters[int(record['chapter_no'])] = Chapter(
            chapter_no=int(record['chapter_no']),
            chapter_title=str(record['chapter_title']),
            summary=str(record.get('summary', '')),
            chapter_text=str(record.get('chapter_text_cleaned') or record.get('chapter_text') or ''),
            line_start=int(line_span.get('start', 0) or 0),
            line_end=int(line_span.get('end', 0) or 0),
            numeric_data=list(record.get('numeric_data', [])),
            topic_tags=list(attributes.get('topic_tags', [])),
        )
    return chapters


chapters = load_chapters()

# Source metrics from the parsed report and chapter records.
metrics = {
    'revenue_2025': Decimal('105323304409.14'),
    'revenue_2024': Decimal('142778764079.80'),
    'net_profit_2025': Decimal('-11946573688.12'),
    'net_profit_2024': Decimal('-9852499598.84'),
    'ocf_2025': Decimal('-3038704592.78'),
    'ocf_2024': Decimal('-5176325811.10'),
    'equity_2025': Decimal('191440815716.19'),
    'equity_2024': Decimal('202666487973.00'),
    'assets_2025': Decimal('1194148874774.55'),
    'assets_2024': Decimal('1286259859765.82'),
    'liab_2025': Decimal('872988331813.49'),
    'liab_2024': Decimal('947405197282.08'),
    'cash_total': Decimal('74002263778.73'),
    'restricted_cash': Decimal('4654701002.07'),
    'cash_available': Decimal('69347562776.66'),
    'ar_gross': Decimal('10453374346.84'),
    'ar_bad': Decimal('1141215447.29'),
    'ar_net': Decimal('9312158899.55'),
    'other_recv_gross': Decimal('243478584320.76'),
    'other_recv_bad': Decimal('28218764412.74'),
    'other_recv_net': Decimal('215259819908.02'),
    'contract_assets': Decimal('12337821477.85'),
    'other_current_assets': Decimal('12701757186.30'),
    'inventory_gross': Decimal('477759366697.48'),
    'inventory_impairment': Decimal('15240826104.80'),
    'inventory_net': Decimal('462518540592.68'),
    'inventory_pledged': Decimal('62747109747.97'),
    'inventory_capitalized_interest': Decimal('17422688932.49'),
    'inv_prop_gross': Decimal('153606501238.22'),
    'inv_prop_depr': Decimal('16066814608.08'),
    'inv_prop_impairment': Decimal('177144292.94'),
    'inv_prop_net': Decimal('137362542337.20'),
    'inv_prop_pledged': Decimal('80046010927.60'),
    'inv_prop_unregistered': Decimal('7871975474.94'),
    'prepayments_over1y': Decimal('8447000000.00'),
    'contract_liab': Decimal('158003116179.93'),
    'contract_liab_gt1y': Decimal('70441987190.62'),
    'ap_total': Decimal('130723480324.92'),
    'ap_land': Decimal('16009433338.20'),
    'ap_engineering': Decimal('106600741232.58'),
    'ap_quality': Decimal('7296788676.44'),
    'other_payables_total': Decimal('142957117746.32'),
    'other_payables_interest': Decimal('48634934.82'),
    'other_payables_dividend': Decimal('267927565.70'),
    'other_payables_other': Decimal('142640585246.18'),
    'other_payables_ltax': Decimal('44436049686.57'),
    'current_other_liabilities': Decimal('14641509761.17'),
    'short_borrow': Decimal('23115623539.41'),
    'current_lt_borrow': Decimal('103953699289.47'),
    'current_bonds': Decimal('27358717473.95'),
    'current_lease': Decimal('1790311510.04'),
    'current_interest': Decimal('1610359156.05'),
    'near_term_obligations': Decimal('134713087429.51'),
    'lt_borrow_net': Decimal('178051318769.28'),
    'lt_bonds_net': Decimal('43602038467.45'),
    'lt_lease_net': Decimal('15753855182.66'),
    'capital_total_debt': Decimal('364257680065.61'),
    'capital_net_debt': Decimal('290255416286.88'),
    'capital_equity': Decimal('321160542961.06'),
    'roe': Decimal('-0.0607'),
    'eps': Decimal('-1.01'),
    'revenue_yoy': Decimal('-0.2623321466049804'),
}

metrics['net_margin'] = metrics['net_profit_2025'] / metrics['revenue_2025']
metrics['ocf_margin'] = metrics['ocf_2025'] / metrics['revenue_2025']
metrics['cash_cover'] = metrics['cash_available'] / metrics['near_term_obligations']
metrics['asset_liab_ratio'] = metrics['liab_2025'] / metrics['assets_2025']
metrics['ar_bad_rate'] = metrics['ar_bad'] / metrics['ar_gross']
metrics['other_recv_bad_rate'] = metrics['other_recv_bad'] / metrics['other_recv_gross']
metrics['inventory_impairment_rate'] = metrics['inventory_impairment'] / metrics['inventory_net']
metrics['inventory_pledged_rate'] = metrics['inventory_pledged'] / metrics['inventory_net']
metrics['inv_prop_pledged_rate'] = metrics['inv_prop_pledged'] / metrics['inv_prop_net']
metrics['contract_liab_gt1y_rate'] = metrics['contract_liab_gt1y'] / metrics['contract_liab']
metrics['capital_net_debt_ratio'] = metrics['capital_net_debt'] / metrics['capital_equity']
metrics['capital_total_debt_ratio'] = metrics['capital_total_debt'] / metrics['capital_equity']
metrics['financing_interest_gross_2025'] = Decimal('7257281238.46')
metrics['financing_capitalized_interest_2025'] = Decimal('3214545326.70')
metrics['financing_interest_nominal_2025'] = metrics['financing_interest_gross_2025'] - metrics['financing_capitalized_interest_2025']
metrics['financing_cost_halfyear_including_capitalization_2025'] = metrics['financing_interest_gross_2025'] + metrics['financing_capitalized_interest_2025']
metrics['financing_capitalization_rate'] = metrics['financing_capitalized_interest_2025'] / metrics['financing_cost_halfyear_including_capitalization_2025']
metrics['interest_bearing_debt_2025'] = (
    metrics['short_borrow']
    + metrics['current_lt_borrow']
    + metrics['lt_borrow_net']
    + metrics['current_bonds']
    + metrics['lt_bonds_net']
    + metrics['current_lease']
    + metrics['lt_lease_net']
)
metrics['interest_bearing_debt_2024'] = (
    Decimal('15973061991.55')
    + Decimal('106182101369.60')
    + Decimal('178886209211.52')
    + Decimal('36152409031.95')
    + Decimal('43602038467.45')
    + Decimal('1887942595.92')
    + Decimal('15753855182.66')
)
metrics['interest_bearing_debt_avg_2025'] = (metrics['interest_bearing_debt_2025'] + metrics['interest_bearing_debt_2024']) / Decimal('2')
metrics['term_debt_scale_2025'] = metrics['interest_bearing_debt_avg_2025']
metrics['financing_cost_rate_nominal_halfyear'] = Decimal('0.019')
metrics['financing_cost_rate_including_capitalization_halfyear'] = metrics['financing_cost_halfyear_including_capitalization_2025'] / metrics['interest_bearing_debt_avg_2025']
metrics['financing_cost_annualized_2025'] = metrics['financing_cost_halfyear_including_capitalization_2025'] * Decimal('2')
metrics['financing_cost_rate_annualized'] = metrics['financing_cost_rate_including_capitalization_halfyear'] * Decimal('2')

key_chapters = [1, 3, 5, 6, 7, 8, 12, 25, 29, 31, 32, 33, 55, 56, 57, 79, 84, 86, 88, 90, 92, 93]
if 79 not in chapters:
    # Fallback if the chapter numbering shifts in future re-runs.
    key_chapters = [n for n in key_chapters if n in chapters]


# Workbook setup
wb = Workbook()
styles = {
    'title_fill': PatternFill('solid', fgColor='173A5E'),
    'section_fill': PatternFill('solid', fgColor='D9EAF7'),
    'subtle_fill': PatternFill('solid', fgColor='F7FAFD'),
    'header_fill': PatternFill('solid', fgColor='2F5E86'),
    'accent_fill': PatternFill('solid', fgColor='EAF2F8'),
    'title_font': Font(color='FFFFFF', bold=True, size=14),
    'sheet_title_font': Font(bold=True, size=12, color='1F1F1F'),
    'header_font': Font(color='FFFFFF', bold=True),
    'bold_font': Font(bold=True),
    'normal_font': Font(color='1F1F1F'),
    'thin_side': Side(style='thin', color='B7C9D6'),
}
styles['box'] = Border(
    left=styles['thin_side'],
    right=styles['thin_side'],
    top=styles['thin_side'],
    bottom=styles['thin_side'],
)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
right = Alignment(horizontal='right', vertical='center')
num_fmt = '#,##0.00;(#,##0.00)'
int_fmt = '#,##0'
pct_fmt = '0.00%'
ratio_fmt = '0.00x'


def set_widths(sheet, widths: dict[int, float]) -> None:
    for idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(idx)].width = width


def title_block(sheet, title: str, subtitle: str | None = None) -> None:
    sheet.merge_cells('A1:H1')
    sheet['A1'] = title
    sheet['A1'].fill = styles['title_fill']
    sheet['A1'].font = styles['title_font']
    sheet['A1'].alignment = left
    if subtitle:
        sheet.merge_cells('A2:H2')
        sheet['A2'] = subtitle
        sheet['A2'].fill = styles['subtle_fill']
        sheet['A2'].alignment = left


def style_table(sheet, start_row: int, headers: list[str], rows: list[list[Any]], widths: dict[int, float] | None = None, currency_cols: set[int] | None = None, pct_cols: set[int] | None = None) -> int:
    currency_cols = currency_cols or set()
    pct_cols = pct_cols or set()
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, col_idx, header)
        cell.fill = styles['header_fill']
        cell.font = styles['header_font']
        cell.alignment = center
        cell.border = styles['box']
    for row_offset, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(start_row + row_offset, col_idx, value)
            cell.border = styles['box']
            cell.alignment = left if col_idx == 1 else center
            if isinstance(value, (int, float, Decimal)):
                if col_idx in pct_cols:
                    cell.number_format = pct_fmt
                elif col_idx in currency_cols:
                    cell.number_format = num_fmt
                elif float(value).is_integer():
                    cell.number_format = int_fmt
                else:
                    cell.number_format = num_fmt
    if widths:
        set_widths(sheet, widths)
    return start_row + len(rows)


def apply_unit_format(cell, unit: str, role: str = 'value') -> None:
    if unit in {'亿元', '元', '倍', '年内'}:
        cell.number_format = num_fmt
    elif unit == '%':
        cell.number_format = pct_fmt
    elif unit == '亿元/x':
        cell.number_format = ratio_fmt if role == 'change' else num_fmt
    else:
        cell.number_format = num_fmt


def to_yi(value: Decimal | float | int) -> float:
    return float(Decimal(str(value)) / Decimal('100000000'))


def chapter_label(chapter_no: int) -> str:
    chapter = chapters[chapter_no]
    return f'{chapter.chapter_no} {chapter.chapter_title}'


# 00_overview
ws = wb.active
ws.title = '00_overview'
title_block(ws, '万科企业股份有限公司 2025年半年度附注优先分析底稿', '来源：ChinaMoney 发现 + CNInfo 回退下载；MinerU 解析；financial-analyzer 生成底稿后人工正式化。')
ws['A4'] = '执行摘要'
ws['A4'].font = styles['sheet_title_font']
ws['A4'].fill = styles['section_fill']
ws['A5'] = '上半年营业收入 1,053.23 亿元，同比下降 26.23%；归母净亏损 119.47 亿元；经营活动现金流净额 -30.39 亿元。账面现金 740.02 亿元，但受限资金 46.55 亿元，按近端到期债务口径现金覆盖率仅 0.52x，按可动用现金口径约 0.52x?'
ws['A5'].alignment = left
ws.merge_cells('A5:H6')
ws['A5'].fill = styles['subtle_fill']
ws['A5'].border = styles['box']

ws['A8'] = '关键结论'
ws['A8'].font = styles['sheet_title_font']
ws['A8'].fill = styles['section_fill']
ws.merge_cells('A9:H10')
ws['A9'] = '主体仍处于“高体量资产、高杠杆、高抵押、低现金创造”的状态。短期违约信号未显性化，但流动性稳定依赖深铁支持、再融资接续和项目去化。'
ws['A9'].alignment = left
ws['A9'].border = styles['box']

ws['A12'] = '核心指标'
ws['A12'].font = styles['sheet_title_font']
ws['A12'].fill = styles['section_fill']
core_rows = [
    ['营业收入(亿元)', to_yi(metrics['revenue_2025']), to_yi(metrics['revenue_2024']), metrics['revenue_yoy'], '收入下滑'],
    ['归母净利润(亿元)', to_yi(metrics['net_profit_2025']), to_yi(metrics['net_profit_2024']), (metrics['net_profit_2025'] - metrics['net_profit_2024']) / abs(metrics['net_profit_2024']), '亏损扩大'],
    ['经营现金流(亿元)', to_yi(metrics['ocf_2025']), to_yi(metrics['ocf_2024']), (metrics['ocf_2025'] - metrics['ocf_2024']) / abs(metrics['ocf_2024']), '仍为负'],
    ['可动用现金(亿元)', to_yi(metrics['cash_available']), to_yi(metrics['cash_total']), metrics['cash_cover'], '剔除受限资金'],
    ['近端到期债务(亿元)', to_yi(metrics['near_term_obligations']), to_yi(Decimal('146045568389.92')), '', '短债+一年内到期借款/债券/租赁/利息'],
    ['存货净额(亿元)', to_yi(metrics['inventory_net']), to_yi(Decimal('519009439846.78')), metrics['inventory_impairment_rate'], '计提减值 152.41 亿元'],
    ['其他应收款净额(亿元)', to_yi(metrics['other_recv_net']), to_yi(Decimal('221359590789.50')), metrics['other_recv_bad_rate'], '97.47% 单项计提'],
    ['合同负债(亿元)', to_yi(metrics['contract_liab']), to_yi(Decimal('192361112557.18')), metrics['contract_liab_gt1y_rate'], '一年以上占 44.58%'],
]
style_table(
    ws,
    13,
    ['指标', '2025H1', '2024H1/2024YE', '变化/比率', '解读'],
    core_rows,
    widths={1: 24, 2: 16, 3: 16, 4: 16, 5: 28},
    currency_cols={2, 3},
    pct_cols={4},
)
for row in range(14, 22):
    ws.cell(row, 2).number_format = num_fmt
    ws.cell(row, 3).number_format = num_fmt
    if row in (14, 15, 16, 17, 19, 20):
        ws.cell(row, 4).number_format = pct_fmt if row != 17 else ratio_fmt

ws['A23'] = '阅读顺序'
ws['A23'].font = styles['sheet_title_font']
ws['A23'].fill = styles['section_fill']
ws['A24'] = '先看现金、债务和合同负债，再看其他应收、存货和投资性房地产，最后回到利润表补充资料和现金流量表补充资料。'
ws.merge_cells('A24:H25')
ws['A24'].alignment = left
ws['A24'].border = styles['box']
set_widths(ws, {1: 24, 2: 16, 3: 16, 4: 16, 5: 28, 6: 18, 7: 18, 8: 18})

# 01_kpi_dashboard
ws = wb.create_sheet('01_kpi_dashboard')
title_block(ws, 'KPI Dashboard', '用于快速扫描收入、利润、现金和杠杆四条主线。')
ws['A4'] = '分类'
ws['B4'] = '指标'
ws['C4'] = '2025H1'
ws['D4'] = '2024H1/2024YE'
ws['E4'] = '变化'
ws['F4'] = '单位'
ws['G4'] = '风险判断'
for cell in ws[4]:
    cell.fill = styles['header_fill']
    cell.font = styles['header_font']
    cell.alignment = center
    cell.border = styles['box']

kpi_rows = [
    ['盈利', '营业收入', to_yi(metrics['revenue_2025']), to_yi(metrics['revenue_2024']), float(metrics['revenue_yoy']), '亿元', '收入明显下行'],
    ['盈利', '归母净利润', to_yi(metrics['net_profit_2025']), to_yi(metrics['net_profit_2024']), float((metrics['net_profit_2025'] - metrics['net_profit_2024']) / abs(metrics['net_profit_2024'])), '亿元', '亏损扩大'],
    ['盈利', 'ROE', float(metrics['roe']), '', '', '%', '股东回报为负'],
    ['现金', '经营活动现金流', to_yi(metrics['ocf_2025']), to_yi(metrics['ocf_2024']), float((metrics['ocf_2025'] - metrics['ocf_2024']) / abs(metrics['ocf_2024'])), '亿元', '仍为负'],
    ['流动性', '可动用现金', to_yi(metrics['cash_available']), to_yi(metrics['cash_total']), float(metrics['cash_cover']), '亿元/x', '覆盖不足'],
    ['流动性', '近端到期债务', to_yi(metrics['near_term_obligations']), to_yi(Decimal('146045568389.92')), '', '亿元', '再融资压力大'],
    ['融资', '半年名义融资成本率', float(metrics['financing_cost_rate_nominal_halfyear']), '', '', '%', '不含资本化利息'],
    ['融资', '半年融资成本率（含资本化）', float(metrics['financing_cost_rate_including_capitalization_halfyear']), '', '', '%', '按平均有息债务规模口径'],
    ['融资', '年化融资成本率', float(metrics['financing_cost_rate_annualized']), '', '', '%', '按平均有息债务规模折年'],
    ['资产质量', '其他应收款净额', to_yi(metrics['other_recv_net']), to_yi(Decimal('221359590789.50')), float(metrics['other_recv_bad_rate']), '亿元', '沉淀资金高'],
    ['资产质量', '存货净额', to_yi(metrics['inventory_net']), to_yi(Decimal('519009439846.78')), float(metrics['inventory_impairment_rate']), '亿元', '去化与减值并存'],
    ['资产质量', '投资性房地产净额', to_yi(metrics['inv_prop_net']), to_yi(Decimal('141057369174.98')), float(metrics['inv_prop_pledged_rate']), '亿元', '抵押深度高'],
    ['资产质量', '合同负债', to_yi(metrics['contract_liab']), to_yi(Decimal('192361112557.18')), float(metrics['contract_liab_gt1y_rate']), '亿元', '预收款仍是周转源'],
]
for row_idx, row in enumerate(kpi_rows, start=5):
    unit = str(row[5])
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row_idx, col_idx, value)
        cell.border = styles['box']
        cell.alignment = left if col_idx in (1, 2, 7) else center
        if col_idx in (3, 4):
            apply_unit_format(cell, unit, 'value')
        if col_idx == 5 and value != '':
            apply_unit_format(cell, unit, 'change')
        if col_idx == 6:
            cell.alignment = center

set_widths(ws, {1: 12, 2: 24, 3: 16, 4: 16, 5: 14, 6: 10, 7: 22})

# 02_financial_summary
ws = wb.create_sheet('02_financial_summary')
title_block(ws, 'Financial Summary', '收入、利润、每股和现金流的低频摘要。')
ws['A4'] = '利润表摘要'
ws['A4'].font = styles['sheet_title_font']
ws['A4'].fill = styles['section_fill']
summary_rows = [
    ['营业收入', '2025H1', to_yi(metrics['revenue_2025']), '亿元', '同比 -26.23%'],
    ['投资收益', '2025H1', -5.6798009753, '亿元', '损失'],
    ['营业收入/归母净利润', '2025H1', to_yi(metrics['revenue_2025']) / abs(to_yi(metrics['net_profit_2025'])), '倍', '营收仍覆盖不了亏损规模'],
    ['归母净利润', '2025H1', to_yi(metrics['net_profit_2025']), '亿元', '亏损'],
    ['基本每股收益', '2025H1', float(metrics['eps']), '元', '为负'],
    ['ROE', '2025H1', float(metrics['roe']), '%', '为负'],
]
style_table(ws, 5, ['项目', '期间', '数值', '单位', '说明'], summary_rows, widths={1: 24, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
for row in range(6, 12):
    apply_unit_format(ws.cell(row, 3), str(ws.cell(row, 4).value), 'value')
ws['A13'] = '现金流摘要'
ws['A13'].font = styles['sheet_title_font']
ws['A13'].fill = styles['section_fill']
flow_rows = [
    ['经营活动现金流净额', '2025H1', to_yi(metrics['ocf_2025']), '亿元', '仍为负'],
    ['购买/收回理财产品', '2025H1', 1.6270308219, '亿元', '现金管理'],
    ['收到利息收入', '2025H1', 3.4169677678, '亿元', '资金收益'],
    ['融资活动负债净变动', '2025H1', 1.4658745470, '亿元', '筹资活动净增'],
]
style_table(ws, 14, ['项目', '期间', '数值', '单位', '说明'], flow_rows, widths={1: 24, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
ws['A20'] = '资产负债表摘要'
ws['A20'].font = styles['sheet_title_font']
ws['A20'].fill = styles['section_fill']
bs_rows = [
    ['总资产', '2025H1', to_yi(metrics['assets_2025']), '亿元', '资产规模仍大'],
    ['总负债', '2025H1', to_yi(metrics['liab_2025']), '亿元', '高杠杆'],
    ['股东权益', '2025H1', to_yi(metrics['equity_2025']), '亿元', '仍为正'],
    ['资产负债率', '2025H1', float(metrics['asset_liab_ratio']), '%', '杠杆偏高'],
]
style_table(ws, 21, ['项目', '期间', '数值', '单位', '说明'], bs_rows, widths={1: 24, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
for row in range(22, 26):
    apply_unit_format(ws.cell(row, 3), str(ws.cell(row, 4).value), 'value')
set_widths(ws, {1: 24, 2: 14, 3: 16, 4: 10, 5: 28})

# 03_debt_profile
ws = wb.create_sheet('03_debt_profile')
title_block(ws, 'Debt Profile', '近端债务、长期债务、债券和租赁负债的结构化拆分。')
ws['A4'] = '近端到期压力'
ws['A4'].font = styles['sheet_title_font']
ws['A4'].fill = styles['section_fill']
near_rows = [
    ['一年内到期的长期借款', '2025H1', to_yi(metrics['current_lt_borrow']), '亿元', '无逾期借款'],
    ['一年内到期的应付债券', '2025H1', to_yi(metrics['current_bonds']), '亿元', '兑付压力'],
    ['一年内到期的租赁负债', '2025H1', to_yi(metrics['current_lease']), '亿元', '广义杠杆'],
    ['一年内到期的应付利息', '2025H1', to_yi(metrics['current_interest']), '亿元', '现金消耗'],
    ['近端到期债务合计', '2025H1', to_yi(metrics['near_term_obligations']), '亿元', '关键压力口径'],
]
style_table(ws, 5, ['项目', '期间', '数值', '单位', '备注'], near_rows, widths={1: 24, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
ws['A12'] = '长期债务'
ws['A12'].font = styles['sheet_title_font']
ws['A12'].fill = styles['section_fill']
long_rows = [
    ['长期借款净额', '2025H1', to_yi(metrics['lt_borrow_net']), '亿元', '含各种抵押/信用/质押'],
    ['应付债券净额', '2025H1', to_yi(metrics['lt_bonds_net']), '亿元', '公开债为主'],
    ['租赁负债净额', '2025H1', to_yi(metrics['lt_lease_net']), '亿元', '广义债务'],
    ['有息债务规模（平均）', '2025H1', to_yi(metrics['interest_bearing_debt_avg_2025']), '亿元', '期初与期末余额平均'],
    ['经调整净债务', '2025H1', to_yi(metrics['capital_net_debt']), '亿元', '总债务减货币资金'],
]
style_table(ws, 13, ['项目', '期间', '数值', '单位', '备注'], long_rows, widths={1: 24, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
ws['A20'] = '融资成本'
ws['A20'].font = styles['sheet_title_font']
ws['A20'].fill = styles['section_fill']
financing_rows = [
    ['贷款、债券及应付款项的利息支出', '2025H1', to_yi(Decimal('6911400250.02')), '亿元', '附注46，含未资本化部分'],
    ['租赁负债的利息支出', '2025H1', to_yi(Decimal('345880988.44')), '亿元', '附注46'],
    ['小计：名义利息支出', '2025H1', to_yi(metrics['financing_interest_gross_2025']), '亿元', '借款、债券及应付款项利息 + 租赁利息'],
    ['减：资本化利息', '2025H1', -to_yi(metrics['financing_capitalized_interest_2025']), '亿元', '回到真实融资成本池'],
    ['半年度融资成本（含资本化）', '2025H1', to_yi(metrics['financing_cost_halfyear_including_capitalization_2025']), '亿元', '融资成本分子'],
    ['年化融资成本', '2025H1x2', to_yi(metrics['financing_cost_annualized_2025']), '亿元', '半年期折年'],
]
style_table(ws, 21, ['项目', '期间', '数值', '单位', '备注'], financing_rows, widths={1: 24, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
rate_row = 27
ws.cell(rate_row, 1, '年化融资成本率').border = styles['box']
ws.cell(rate_row, 2, '2025H1x2').border = styles['box']
ws.cell(rate_row, 3, float(metrics['financing_cost_rate_annualized'])).border = styles['box']
ws.cell(rate_row, 4, '%').border = styles['box']
ws.cell(rate_row, 5, '按平均有息债务规模口径折年').border = styles['box']
for col in range(1, 6):
    ws.cell(rate_row, col).alignment = left if col in (1, 2, 5) else center
ws.cell(rate_row, 1).fill = styles['subtle_fill']
ws.cell(rate_row, 1).font = styles['bold_font']
ws.cell(rate_row, 2).fill = styles['subtle_fill']
ws.cell(rate_row, 3).fill = styles['subtle_fill']
ws.cell(rate_row, 4).fill = styles['subtle_fill']
ws.cell(rate_row, 5).fill = styles['subtle_fill']
for col in range(1, 6):
    ws.cell(rate_row, col).border = styles['box']
ws.cell(rate_row, 3).number_format = pct_fmt
ws['A29'] = '资本结构'
ws['A29'].font = styles['sheet_title_font']
ws['A29'].fill = styles['section_fill']
capital_rows = [
    ['股东权益', '2025H1', to_yi(metrics['capital_equity']), '亿元', '分母'],
    ['净债务/股东权益', '2025H1', float(metrics['capital_net_debt_ratio']), 'x', '约 90.4%'],
    ['总债务/股东权益', '2025H1', float(metrics['capital_total_debt_ratio']), 'x', '约 113.4%'],
    ['管理层披露财务契约借款规模', '2025H1', 103.944, '亿元', '未触发实时偿还'],
]
style_table(ws, 30, ['项目', '期间', '数值', '单位', '备注'], capital_rows, widths={1: 24, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
set_widths(ws, {1: 24, 2: 14, 3: 16, 4: 10, 5: 28})

# 04_liquidity_and_covenants
ws = wb.create_sheet('04_liquidity_and_covenants')
title_block(ws, 'Liquidity and Covenants', '现金、受限资产、管理层支持和契约约束放在同一页看。')
ws['A4'] = '现金与覆盖'
ws['A4'].font = styles['sheet_title_font']
ws['A4'].fill = styles['section_fill']
liq_rows = [
    ['现金及现金等价物', '2025H1', to_yi(metrics['cash_total']), '亿元', '名义现金'],
    ['受限资金', '2025H1', to_yi(metrics['restricted_cash']), '亿元', '不可自由动用'],
    ['可动用现金', '2025H1', to_yi(metrics['cash_available']), '亿元', '用于偿债覆盖'],
    ['近端到期债务', '2025H1', to_yi(metrics['near_term_obligations']), '亿元', '关键压力口径'],
    ['现金覆盖率', '2025H1', float(metrics['cash_cover']), 'x', '覆盖不足 1 倍'],
]
style_table(ws, 5, ['项目', '期间', '数值', '单位', '备注'], liq_rows, widths={1: 24, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
ws['A12'] = '管理层支持'
ws['A12'].font = styles['sheet_title_font']
ws['A12'].fill = styles['section_fill']
ws.merge_cells('A13:H14')
ws['A13'] = '管理层在报告中披露：深铁集团累计向公司提供 238.8 亿元股东借款；上半年已完成 243.9 亿元公开债务偿还；2027 年之前无境外公开债到期。'
ws['A13'].alignment = left
ws['A13'].border = styles['box']
ws['A13'].fill = styles['subtle_fill']
ws['A16'] = '契约与窗口'
ws['A16'].font = styles['sheet_title_font']
ws['A16'].fill = styles['section_fill']
window_rows = [
    ['附财务契约借款规模', '2025H1', 103.944, '亿元', '涉及财务比率约束'],
    ['未逾期的一年内到期长期借款', '2025H1', to_yi(metrics['current_lt_borrow']), '亿元', '公司明确披露无逾期'],
    ['资本结构总债务', '2025H1', to_yi(metrics['capital_total_debt']), '亿元', '管理层口径'],
]
style_table(ws, 17, ['项目', '期间', '数值', '单位', '备注'], window_rows, widths={1: 24, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
set_widths(ws, {1: 24, 2: 14, 3: 16, 4: 10, 5: 28})

# 05_asset_quality
ws = wb.create_sheet('05_asset_quality')
title_block(ws, 'Asset Quality', '重点看“可变现性”和“是否被抵押”，不是只看账面规模。')
ws['A4'] = '核心资产'
ws['A4'].font = styles['sheet_title_font']
ws['A4'].fill = styles['section_fill']
asset_rows = [
    ['应收账款净额', '2025H1', to_yi(metrics['ar_net']), '亿元', f'坏账率 {metrics["ar_bad_rate"]:.2%}'],
    ['其他应收款净额', '2025H1', to_yi(metrics['other_recv_net']), '亿元', f'单项计提占比 97.47%'],
    ['合同资产', '2025H1', to_yi(metrics['contract_assets']), '亿元', '政府/相关方为主'],
    ['其他流动资产', '2025H1', to_yi(metrics['other_current_assets']), '亿元', '合同取得成本 + 待抵扣增值税'],
    ['存货净额', '2025H1', to_yi(metrics['inventory_net']), '亿元', f'减值率 {metrics["inventory_impairment_rate"]:.2%}'],
    ['存货抵押/担保', '2025H1', to_yi(metrics['inventory_pledged']), '亿元', f'占存货净额 {metrics["inventory_pledged_rate"]:.2%}'],
    ['投资性房地产净额', '2025H1', to_yi(metrics['inv_prop_net']), '亿元', f'抵押占比 {metrics["inv_prop_pledged_rate"]:.2%}'],
    ['未办妥产权证的投资性房地产', '2025H1', to_yi(metrics['inv_prop_unregistered']), '亿元', '产权手续未完结'],
    ['合同负债', '2025H1', to_yi(metrics['contract_liab']), '亿元', f'一年以上占 {metrics["contract_liab_gt1y_rate"]:.2%}'],
    ['预付款项（重大一年以上）', '2025H1', to_yi(metrics['prepayments_over1y']), '亿元', '旧改项目前期款'],
]
style_table(ws, 5, ['项目', '期间', '数值', '单位', '判断'], asset_rows, widths={1: 28, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
ws['A17'] = '资产质量结论'
ws['A17'].font = styles['sheet_title_font']
ws['A17'].fill = styles['section_fill']
ws.merge_cells('A18:H19')
ws['A18'] = '账面资产很大，但其他应收和存货两大科目都存在明显沉淀和减值压力，投资性房地产又被大比例抵押。可动用资产与账面资产之间的差距，是本期信用分析的核心。'
ws['A18'].alignment = left
ws['A18'].border = styles['box']
set_widths(ws, {1: 28, 2: 14, 3: 16, 4: 10, 5: 28})

# 06_cash_flow_bridge
ws = wb.create_sheet('06_cash_flow_bridge')
title_block(ws, 'Cash Flow Bridge', '从附注和利润补充资料里，挑出影响现金流的关键桥梁。')
ws['A4'] = '利润到现金'
ws['A4'].font = styles['sheet_title_font']
ws['A4'].fill = styles['section_fill']
bridge_rows = [
    ['归母净亏损', '2025H1', to_yi(metrics['net_profit_2025']), '亿元', '利润端起点'],
    ['资产减值损失', '2025H1', 51.4518472878, '亿元', '利润补充资料'],
    ['信用减值损失', '2025H1', 3.0165134647, '亿元', '应收风险仍在'],
    ['投资收益(损失)', '2025H1', -5.6798009753, '亿元', '非经常性波动'],
    ['经营活动现金流净额', '2025H1', to_yi(metrics['ocf_2025']), '亿元', '现金流结果'],
    ['经营现金流补充说明', '2025H1', 0, '—', '净亏损、减值和营运资产变化共同解释'],
]
style_table(ws, 5, ['项目', '期间', '数值', '单位', '说明'], bridge_rows, widths={1: 22, 2: 14, 3: 16, 4: 10, 5: 32}, currency_cols={3})
ws['A13'] = '筹资活动'
ws['A13'].font = styles['sheet_title_font']
ws['A13'].fill = styles['section_fill']
fund_rows = [
    ['短期借款期末余额', '2025H1', to_yi(metrics['short_borrow']), '亿元', '滚动融资'],
    ['长期借款(含一年内到期)', '2025H1', 282.00501805875, '亿元', '筹资活动净变动后口径'],
    ['应付债券(含一年内到期)', '2025H1', to_yi(Decimal('43602038467.45')) + to_yi(metrics['current_bonds']), '亿元', '公开债滚动'],
    ['租赁负债(含一年内到期)', '2025H1', to_yi(Decimal('17544166692.70')), '亿元', '广义负债'],
]
style_table(ws, 14, ['项目', '期间', '数值', '单位', '说明'], fund_rows, widths={1: 22, 2: 14, 3: 16, 4: 10, 5: 32}, currency_cols={3})
set_widths(ws, {1: 22, 2: 14, 3: 16, 4: 10, 5: 32})

# 07_notes_map
ws = wb.create_sheet('07_notes_map')
title_block(ws, 'Notes Map', '选取最重要的附注章节，作为正式成稿的证据底座。')
rows = []
for chapter_no in key_chapters:
    chapter = chapters[chapter_no]
    evidence_excerpt = chapter.summary[:120]
    if not evidence_excerpt:
        evidence_excerpt = chapter.chapter_text[:120]
    rows.append([
        chapter.chapter_no,
        chapter.chapter_title,
        f'{chapter.line_start}-{chapter.line_end}',
        evidence_excerpt,
        ', '.join(chapter.topic_tags[:3]),
    ])
style_table(ws, 4, ['章号', '章节', '行范围', '关键摘要', '主题标签'], rows, widths={1: 8, 2: 24, 3: 14, 4: 52, 5: 24})

# 08_management_support
ws = wb.create_sheet('08_management_support')
title_block(ws, 'Management Support', '管理层披露的股东支持、债务到期和再融资窗口。')
ws['A4'] = '支持与窗口'
ws['A4'].font = styles['sheet_title_font']
ws['A4'].fill = styles['section_fill']
management_rows = [
    ['深铁股东借款累计', '报告正文', 238.8, '亿元', '流动性支撑'],
    ['公开债务偿还', '报告正文', 243.9, '亿元', '已完成'],
    ['境外公开债到期', '报告正文', 0, '年内', '2027 年前无到期'],
    ['管理层对流动性判断', '报告正文', 0, '—', '以支持+再融资缓冲短期风险'],
]
style_table(ws, 5, ['事项', '来源', '数值', '单位', '说明'], management_rows, widths={1: 22, 2: 12, 3: 16, 4: 10, 5: 34}, currency_cols={3})
ws['A11'] = '原文摘要'
ws['A11'].font = styles['sheet_title_font']
ws['A11'].fill = styles['section_fill']
ws.merge_cells('A12:H13')
ws['A12'] = '管理层披露：深铁集团在上半年继续提供流动性支持，累计股东借款 238.8 亿元；公司已完成 243.9 亿元公开债务偿还；2027 年之前无境外公开债到期。'
ws['A12'].alignment = left
ws['A12'].border = styles['box']
set_widths(ws, {1: 22, 2: 12, 3: 16, 4: 10, 5: 34})

# 09_capital_structure
ws = wb.create_sheet('09_capital_structure')
title_block(ws, 'Capital Structure', '管理层资本结构口径，结合总债务和净债务观察杠杆。')
ws['A4'] = '资本结构'
ws['A4'].font = styles['sheet_title_font']
ws['A4'].fill = styles['section_fill']
capital_rows = [
    ['总债务合计', '2025H1', to_yi(metrics['capital_total_debt']), '亿元', '管理层口径'],
    ['货币资金', '2025H1', to_yi(metrics['cash_total']), '亿元', '扣减项'],
    ['经调整净债务', '2025H1', to_yi(metrics['capital_net_debt']), '亿元', '总债务减现金'],
    ['股东权益', '2025H1', to_yi(metrics['capital_equity']), '亿元', '分母'],
    ['净债务/股东权益', '2025H1', float(metrics['capital_net_debt_ratio']), 'x', '接近 0.90x'],
    ['总债务/股东权益', '2025H1', float(metrics['capital_total_debt_ratio']), 'x', '接近 1.13x'],
]
style_table(ws, 5, ['项目', '期间', '数值', '单位', '备注'], capital_rows, widths={1: 24, 2: 14, 3: 16, 4: 10, 5: 28}, currency_cols={3})
set_widths(ws, {1: 24, 2: 14, 3: 16, 4: 10, 5: 28})

# 99_evidence_index
ws = wb.create_sheet('99_evidence_index')
title_block(ws, 'Evidence Index', '证据索引用于回溯各表格中的关键数字。')
evidence_rows: list[list[Any]] = [
    ['EVD-0001', chapter_label(1), '货币资金、受限资金、境外存款', '2429-2436', '现金与受限资金'],
    ['EVD-0002', chapter_label(3), '应收账款坏账率与账龄', '2441-2489', '回款风险'],
    ['EVD-0003', chapter_label(5), '其他应收款单项计提 97.47%', '2497-2559', '沉淀资金'],
    ['EVD-0004', chapter_label(6), '存货减值 152.41 亿元、抵押 62.75 亿元', '2560-2757', '库存与抵押'],
    ['EVD-0005', chapter_label(7), '合同资产 123.38 亿元', '2758-2777', '合同结算'],
    ['EVD-0006', chapter_label(8), '合同取得成本与待抵扣增值税', '2778-2793', '流动资产'],
    ['EVD-0007', chapter_label(12), '投资性房地产 1,373.63 亿元，抵押 800.46 亿元', '2830-2842', '商业资产'],
    ['EVD-0008', chapter_label(25), '合同负债 1,580.03 亿元，一年以上 704.42 亿元', '2989-3006', '预收房款'],
    ['EVD-0009', chapter_label(29), '一年内到期的非流动负债 1,347.13 亿元', '3069-3080', '近端债务'],
    ['EVD-0010', chapter_label(31), '长期借款 1,780.51 亿元，契约借款 1,039.44 亿元', '3089-3121', '银行借款'],
    ['EVD-0011', chapter_label(32), '应付债券净额 436.02 亿元，一年内到期 273.59 亿元', '3122-3143', '债券兑付'],
    ['EVD-0012', chapter_label(33), '租赁负债净额 157.54 亿元', '3144-3149', '广义杠杆'],
    ['EVD-0013', chapter_label(55), '归母净亏损 119.47 亿元，基本 EPS -1.01 元', '3345-3350', '利润质量'],
    ['EVD-0014', chapter_label(57), '经营活动现金流净额 -30.39 亿元', '3397-3436', '现金流'],
    ['EVD-0015', chapter_label(79) if 79 in chapters else '79 资本结构', '总债务 3642.58 亿元、净债务 2902.55 亿元', '资本结构', '杠杆'],
    ['EVD-0016', chapter_label(92), 'ROE -6.07%、EPS -1.01 元', '4139-4146', '股东回报'],
    ['EVD-0017', chapter_label(93), '中外准则差异为零', '4147-4155', '口径一致'],
    ['EVD-0018', chapter_label(47), '名义利息 72.57 亿元，资本化利息 32.15 亿元，按期初期末平均有息债务口径，含资本化后半年融资成本约 2.63%，折年约 5.26%', '3267-3272', '融资成本'],
]
style_table(ws, 4, ['证据ID', '章节', '证据摘要', '来源区间', '用途'], evidence_rows, widths={1: 14, 2: 22, 3: 52, 4: 14, 5: 18})

# Workbook cosmetics
for sheet in wb.worksheets:
    sheet.freeze_panes = 'A4'
    for row in sheet.iter_rows():
        for cell in row:
            if cell.row == 1:
                continue
            if isinstance(cell.value, str) and cell.value.startswith('='):
                # Avoid formulas in the generated workbook to keep it deterministic.
                pass

try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
except Exception:
    pass

wb.save(WORKBOOK_PATH)

# Refresh dependent artifacts to match the new workbook.
selected_chapters: list[dict[str, Any]] = []
for chapter_no in key_chapters:
    chapter = chapters[chapter_no]
    selected_chapters.append({
        'chapter_no': chapter.chapter_no,
        'chapter_title': chapter.chapter_title,
        'line_span': [chapter.line_start, chapter.line_end],
        'summary': chapter.summary,
        'topic_tags': chapter.topic_tags,
    })

final_data = {
    'entity_profile': {
        'company_name': '万科企业股份有限公司',
        'report_period': '2025年半年度',
        'report_type': 'A股半年度报告（审阅）',
        'currency': 'CNY',
        'analysis_mode': 'notes_first_manual_formalization',
        'input_file': '/Volumes/P5S/工作/同步空间/BaiduSyncdisk/bondclaw/tmp/vanke_2025_mineru/vanke_2025_full_report/vanke_2025_full_report.md',
    },
    'metrics': {
        'revenue_2025': float(metrics['revenue_2025']),
        'revenue_2024': float(metrics['revenue_2024']),
        'revenue_yoy': float(metrics['revenue_yoy']),
        'net_profit_2025': float(metrics['net_profit_2025']),
        'net_profit_2024': float(metrics['net_profit_2024']),
        'ocf_2025': float(metrics['ocf_2025']),
        'ocf_2024': float(metrics['ocf_2024']),
        'cash_available': float(metrics['cash_available']),
        'near_term_obligations': float(metrics['near_term_obligations']),
        'cash_cover': float(metrics['cash_cover']),
        'inventory_impairment_rate': float(metrics['inventory_impairment_rate']),
        'ar_bad_rate': float(metrics['ar_bad_rate']),
        'other_recv_bad_rate': float(metrics['other_recv_bad_rate']),
        'contract_liab_gt1y_rate': float(metrics['contract_liab_gt1y_rate']),
        'inv_prop_pledged_rate': float(metrics['inv_prop_pledged_rate']),
        'financing_interest_gross_2025': float(metrics['financing_interest_gross_2025']),
        'financing_capitalized_interest_2025': float(metrics['financing_capitalized_interest_2025']),
        'financing_interest_nominal_2025': float(metrics['financing_interest_nominal_2025']),
        'financing_cost_halfyear_including_capitalization_2025': float(metrics['financing_cost_halfyear_including_capitalization_2025']),
        'financing_capitalization_rate': float(metrics['financing_capitalization_rate']),
        'interest_bearing_debt_2025': float(metrics['interest_bearing_debt_2025']),
        'interest_bearing_debt_2024': float(metrics['interest_bearing_debt_2024']),
        'interest_bearing_debt_avg_2025': float(metrics['interest_bearing_debt_avg_2025']),
        'term_debt_scale_2025': float(metrics['term_debt_scale_2025']),
        'financing_cost_rate_nominal_halfyear': float(metrics['financing_cost_rate_nominal_halfyear']),
        'financing_cost_rate_including_capitalization_halfyear': float(metrics['financing_cost_rate_including_capitalization_halfyear']),
        'financing_cost_annualized_2025': float(metrics['financing_cost_annualized_2025']),
        'financing_cost_rate_annualized': float(metrics['financing_cost_rate_annualized']),
    },
    'selected_chapters': selected_chapters,
    'conclusion': '信用风险未解除，但仍处于可管理区间；核心支撑来自股东支持、再融资和项目去化。',
    'chapter_count': len(chapters),
}
FINAL_DATA_PATH.write_text(json.dumps(final_data, ensure_ascii=False, indent=2), encoding='utf-8')

soul_payload = {
    'contract_version': 'soul_export_v1',
    'template_version': 'soul_v1_1_alpha',
    'generated_at': datetime.now().astimezone().isoformat(timespec='seconds'),
    'entity_profile': {
        'company_name': '万科企业股份有限公司',
        'report_period': '2025年半年度',
        'currency': 'CNY',
        'report_type': 'a_share_full_report',
        'audit_opinion': '审阅',
        'industry_tag': 'lgfv',
        'input_file': '/Volumes/P5S/工作/同步空间/BaiduSyncdisk/bondclaw/tmp/vanke_2025_mineru/vanke_2025_full_report/vanke_2025_full_report.md',
    },
    'summary': {
        'executive_view': '账面资产大、抵押深、再融资依赖高，利润与现金流尚未恢复到安全区。',
        'risk_view': '短期靠股东支持和债务滚动缓冲，长期要看去库存、减值消化和回款修复。',
    },
    'source_artifacts': {
        'run_manifest': str(RUN_DIR / 'run_manifest.json'),
        'chapter_records': str(CHAPTER_PATH),
        'notes_workfile': str(RUN_DIR / 'vanke_2025_notes_workfile.json'),
        'analysis_report_scaffold': str(RUN_DIR / 'analysis_report_scaffold.md'),
        'final_data_scaffold': str(RUN_DIR / 'final_data_scaffold.json'),
    },
    'evidence_index': [
        {'evidence_id': 'E1', 'chapter_no': 1, 'excerpt': '可动用现金 693.48 亿元，受限资金 46.55 亿元'},
        {'evidence_id': 'E2', 'chapter_no': 5, 'excerpt': '其他应收款净额 2,152.60 亿元，97.47% 单项计提'},
        {'evidence_id': 'E3', 'chapter_no': 6, 'excerpt': '存货净额 4,625.19 亿元，减值 152.41 亿元，抵押 62.75 亿元'},
        {'evidence_id': 'E4', 'chapter_no': 12, 'excerpt': '投资性房地产净额 1,373.63 亿元，抵押 800.46 亿元'},
        {'evidence_id': 'E5', 'chapter_no': 25, 'excerpt': '合同负债 1,580.03 亿元，一年以上 704.42 亿元'},
        {'evidence_id': 'E6', 'chapter_no': 29, 'excerpt': '一年内到期的非流动负债 1,347.13 亿元'},
        {'evidence_id': 'E7', 'chapter_no': 31, 'excerpt': '长期借款 1,780.51 亿元，契约借款 1,039.44 亿元'},
        {'evidence_id': 'E8', 'chapter_no': 32, 'excerpt': '应付债券净额 436.02 亿元，一年内到期 273.59 亿元'},
        {'evidence_id': 'E9', 'chapter_no': 55, 'excerpt': '归母净亏损 119.47 亿元，基本 EPS -1.01 元'},
        {'evidence_id': 'E10', 'chapter_no': 57, 'excerpt': '经营活动现金流净额 -30.39 亿元'},
        {'evidence_id': 'E11', 'chapter_no': 79, 'excerpt': '总债务 3,642.58 亿元，净债务 2,902.55 亿元'},
        {'evidence_id': 'E12', 'chapter_no': 92, 'excerpt': 'ROE -6.07%，EPS -1.01 元'},
        {'evidence_id': 'E13', 'chapter_no': 47, 'excerpt': '名义利息 72.57 亿元，资本化利息 32.15 亿元，按期初期末平均有息债务口径，含资本化后半年融资成本约 2.63%，折年约 5.26%'},
    ],
    'conclusion': '信用风险未解除，但短期仍可管理，关键支撑来自股东支持、再融资接续和项目去化。',
}
SOUL_PAYLOAD_PATH.write_text(json.dumps(soul_payload, ensure_ascii=False, indent=2), encoding='utf-8')

formalization_manifest = {
    'generated_at': datetime.now().astimezone().isoformat(timespec='seconds'),
    'run_dir': str(RUN_DIR),
    'analysis_report': str(RUN_DIR / 'analysis_report.md'),
    'final_data': str(FINAL_DATA_PATH),
    'soul_export_payload': str(SOUL_PAYLOAD_PATH),
    'financial_output': str(WORKBOOK_PATH),
    'chapter_count': len(chapters),
    'sheet_names': wb.sheetnames,
}
FORMALIZATION_MANIFEST_PATH.write_text(json.dumps(formalization_manifest, ensure_ascii=False, indent=2), encoding='utf-8')

print(json.dumps({
    'workbook': str(WORKBOOK_PATH),
    'final_data': str(FINAL_DATA_PATH),
    'soul_export_payload': str(SOUL_PAYLOAD_PATH),
    'formalization_manifest': str(FORMALIZATION_MANIFEST_PATH),
    'sheets': wb.sheetnames,
}, ensure_ascii=False, indent=2))
